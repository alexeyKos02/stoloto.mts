"""
Сравнение текущих файлов с последними бекапами на Яндекс Диске.

Автоматически находит последний *_backup_* файл рядом с оригиналом,
скачивает оба и сравнивает все листы по-ячеечно.

Результат — .xlsx файл-отчёт на Яндекс Диске:
  Для каждой изменённой строки: сверху старая (красные ячейки), снизу новая (зелёные ячейки).
  Удалённые строки — полностью красные, добавленные — полностью зелёные.

ENV:
  YANDEX_OAUTH_TOKEN
  DISK_SOURCE_PATH  — путь к SOURCE на Яндекс Диске
  DISK_TARGET_PATH  — путь к TARGET на Яндекс Диске
"""

import io
import os
import re
import sys
from datetime import datetime, timezone
from typing import Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from sync_utils import disk_download, disk_upload, header_index_map, get_last_data_row


# =======================
# ENV
# =======================
YANDEX_OAUTH_TOKEN = os.getenv("YANDEX_OAUTH_TOKEN", "").strip()
DISK_SOURCE_PATH = os.getenv("DISK_SOURCE_PATH", "").strip()
DISK_TARGET_PATH = os.getenv("DISK_TARGET_PATH", "").strip()

if not YANDEX_OAUTH_TOKEN:
    raise RuntimeError("ERROR: YANDEX_OAUTH_TOKEN is empty")
if not DISK_SOURCE_PATH:
    raise RuntimeError("ERROR: DISK_SOURCE_PATH is empty")
if not DISK_TARGET_PATH:
    raise RuntimeError("ERROR: DISK_TARGET_PATH is empty")

YANDEX_API = "https://cloud-api.yandex.net/v1/disk"

# Стили
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_SECTION = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_SECTION = Font(bold=True, size=11)
FONT_LABEL = Font(bold=True, italic=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# =======================
# FIND LATEST BACKUP
# =======================
def find_latest_backup(original_path: str, token: str) -> Optional[str]:
    folder = os.path.dirname(original_path)
    basename = os.path.basename(original_path)
    name_no_ext, ext = os.path.splitext(basename)

    backup_pattern = re.compile(
        re.escape(name_no_ext) + r'_backup_(\d{4}-\d{2}-\d{2}T\d{2}-\d{2}-\d{2})' + re.escape(ext)
    )

    headers = {"Authorization": f"OAuth {token}"}
    r = requests.get(
        f"{YANDEX_API}/resources",
        headers=headers,
        params={"path": folder, "limit": 200, "fields": "_embedded.items.name,_embedded.items.path"},
        timeout=60,
    )
    if r.status_code >= 400:
        raise RuntimeError(f"LIST ERROR: {r.status_code}\nPATH: {folder}\nBODY: {r.text}")

    items = r.json().get("_embedded", {}).get("items", [])

    backups: List[Tuple[str, str]] = []
    for item in items:
        name = item.get("name", "")
        m = backup_pattern.match(name)
        if m:
            backups.append((m.group(1), item.get("path", "")))

    if not backups:
        return None

    backups.sort(key=lambda x: x[0], reverse=True)
    return backups[0][1]


# =======================
# HELPERS
# =======================
def cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def read_sheet_data(ws: Worksheet) -> Tuple[List[str], List[Dict[str, str]]]:
    """Возвращает (headers, rows). Каждый row — dict {col_name: value_str}."""
    hmap = header_index_map(ws)
    if not hmap:
        return [], []

    headers = list(hmap.keys())
    # Последняя строка по ВСЕМ колонкам (не только первой)
    last = max(get_last_data_row(ws, col_idx, start_row=2) for col_idx in hmap.values())

    rows: List[Dict[str, str]] = []
    for r in range(2, last + 1):
        row_data: Dict[str, str] = {}
        for name, col_idx in hmap.items():
            row_data[name] = cell_str(ws.cell(row=r, column=col_idx).value)
        rows.append(row_data)

    return headers, rows


def row_label(row_data: Dict[str, str]) -> str:
    for key in ["ЮЛ", "Агент ID (Столото)", "Terminal ID (Столото)"]:
        if key in row_data and row_data[key]:
            return f'{key}="{row_data[key]}"'
    for k, v in row_data.items():
        if v:
            return f'{k}="{v}"'
    return ""


# =======================
# WRITE DIFF TO XLSX SHEET
# =======================
def write_diff_sheet(
    ws_report: Worksheet,
    sheet_name: str,
    ws_old: Worksheet,
    ws_new: Worksheet,
    start_row: int,
) -> Tuple[int, int]:
    """
    Записывает diff в ws_report начиная с start_row.
    Возвращает (next_row, change_count).
    """
    headers_old, rows_old = read_sheet_data(ws_old)
    headers_new, rows_new = read_sheet_data(ws_new)

    all_cols = list(dict.fromkeys(headers_old + headers_new))
    if not all_cols:
        return start_row, 0

    changes = 0
    row = start_row

    # Секция: имя листа
    ws_report.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(all_cols) + 1)
    sec_cell = ws_report.cell(row=row, column=1, value=f"Лист: {sheet_name}")
    sec_cell.fill = FILL_SECTION
    sec_cell.font = FONT_SECTION
    row += 1

    # Заголовки таблицы: "Тип" + колонки
    ws_report.cell(row=row, column=1, value="Строка")
    for ci, col_name in enumerate(all_cols):
        c = ws_report.cell(row=row, column=ci + 2, value=col_name)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.border = THIN_BORDER
    h_cell = ws_report.cell(row=row, column=1)
    h_cell.fill = FILL_HEADER
    h_cell.font = FONT_HEADER
    h_cell.border = THIN_BORDER
    row += 1

    max_rows = max(len(rows_old), len(rows_new))
    prev_type = None  # "modified", "deleted", "added"

    for i in range(max_rows):
        excel_row = i + 2  # номер строки в оригинальном файле
        has_old = i < len(rows_old)
        has_new = i < len(rows_new)

        if has_old and has_new:
            old_row = rows_old[i]
            new_row = rows_new[i]

            # Найти изменённые колонки
            changed_cols = set()
            for col in all_cols:
                if old_row.get(col, "") != new_row.get(col, ""):
                    changed_cols.add(col)

            if not changed_cols:
                continue  # строка не изменилась

            if prev_type is not None:
                row += 1
            prev_type = "modified"

            changes += len(changed_cols)

            # Старая строка (красная подсветка изменённых)
            label_cell = ws_report.cell(row=row, column=1, value=f"#{excel_row} БЫЛО")
            label_cell.font = FONT_LABEL
            label_cell.border = THIN_BORDER
            for ci, col in enumerate(all_cols):
                c = ws_report.cell(row=row, column=ci + 2, value=old_row.get(col, ""))
                c.border = THIN_BORDER
                if col in changed_cols:
                    c.fill = FILL_RED
            row += 1

            # Новая строка (зелёная подсветка изменённых)
            label_cell = ws_report.cell(row=row, column=1, value=f"#{excel_row} СТАЛО")
            label_cell.font = FONT_LABEL
            label_cell.border = THIN_BORDER
            for ci, col in enumerate(all_cols):
                c = ws_report.cell(row=row, column=ci + 2, value=new_row.get(col, ""))
                c.border = THIN_BORDER
                if col in changed_cols:
                    c.fill = FILL_GREEN
            row += 1

        elif has_old and not has_new:
            cur_type = "deleted"
            if prev_type is not None and prev_type != cur_type:
                row += 1
            prev_type = cur_type

            # Удалённая строка — вся красная
            changes += 1
            label_cell = ws_report.cell(row=row, column=1, value=f"#{excel_row} УДАЛЕНА")
            label_cell.font = FONT_LABEL
            label_cell.fill = FILL_RED
            label_cell.border = THIN_BORDER
            for ci, col in enumerate(all_cols):
                c = ws_report.cell(row=row, column=ci + 2, value=rows_old[i].get(col, ""))
                c.fill = FILL_RED
                c.border = THIN_BORDER
            row += 1

        elif not has_old and has_new:
            cur_type = "added"
            if prev_type is not None and prev_type != cur_type:
                row += 1
            prev_type = cur_type

            # Добавленная строка — вся зелёная
            changes += 1
            label_cell = ws_report.cell(row=row, column=1, value=f"#{excel_row} ДОБАВЛЕНА")
            label_cell.font = FONT_LABEL
            label_cell.fill = FILL_GREEN
            label_cell.border = THIN_BORDER
            for ci, col in enumerate(all_cols):
                c = ws_report.cell(row=row, column=ci + 2, value=rows_new[i].get(col, ""))
                c.fill = FILL_GREEN
                c.border = THIN_BORDER
            row += 1

    if changes == 0:
        ws_report.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(all_cols) + 1)
        ws_report.cell(row=row, column=1, value="Изменений нет")
        row += 1

    row += 1  # пустая строка между листами
    return row, changes


# =======================
# DIFF WORKBOOKS → REPORT SHEET
# =======================
def diff_to_report_sheet(
    ws_report: Worksheet,
    label: str,
    backup_path: str,
    current_path: str,
    token: str,
    start_row: int,
) -> Tuple[int, int]:
    """Сравнивает два файла, пишет результат в ws_report. Возвращает (next_row, total_changes)."""
    # Заголовок секции
    ws_report.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
    title = ws_report.cell(row=start_row, column=1, value=label)
    title.font = Font(bold=True, size=14)
    start_row += 1

    ws_report.cell(row=start_row, column=1, value=f"Бекап: {os.path.basename(backup_path)}")
    start_row += 1
    ws_report.cell(row=start_row, column=1, value=f"Текущий: {os.path.basename(current_path)}")
    start_row += 2

    old_bytes = disk_download(backup_path, token)
    new_bytes = disk_download(current_path, token)

    wb_old = load_workbook(io.BytesIO(old_bytes))
    wb_new = load_workbook(io.BytesIO(new_bytes))

    all_sheets = list(dict.fromkeys(wb_old.sheetnames + wb_new.sheetnames))
    total = 0
    row = start_row

    for sn in all_sheets:
        if sn in wb_old.sheetnames and sn not in wb_new.sheetnames:
            ws_report.cell(row=row, column=1, value=f"[{sn}] ЛИСТ УДАЛЁН")
            ws_report.cell(row=row, column=1).fill = FILL_RED
            row += 2
            total += 1
            continue
        if sn not in wb_old.sheetnames and sn in wb_new.sheetnames:
            ws_report.cell(row=row, column=1, value=f"[{sn}] ЛИСТ ДОБАВЛЕН")
            ws_report.cell(row=row, column=1).fill = FILL_GREEN
            row += 2
            total += 1
            continue

        row, ch = write_diff_sheet(ws_report, sn, wb_old[sn], wb_new[sn], row)
        total += ch

    return row, total


# =======================
# ENTRYPOINT
# =======================
def main() -> None:
    wb_report = Workbook()
    ws = wb_report.active
    ws.title = "Diff Report"

    # Заголовок
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    ws.merge_cells("A1:J1")
    title = ws.cell(row=1, column=1, value=f"Отчёт сравнения — {ts}")
    title.font = Font(bold=True, size=16)
    row = 3

    total_all = 0

    for label, path in [("SOURCE (ВНУТРЯНКА)", DISK_SOURCE_PATH), ("TARGET (ВНЕШНЯЯ)", DISK_TARGET_PATH)]:
        print(f"Ищу последний бекап для {label}...")
        backup = find_latest_backup(path, YANDEX_OAUTH_TOKEN)

        if not backup:
            print(f"  Бекап не найден — пропускаю")
            ws.cell(row=row, column=1, value=f"{label}: бекап не найден")
            row += 2
            continue

        print(f"  Найден: {backup}")
        print(f"  Сравниваю...")
        row, ch = diff_to_report_sheet(ws, label, backup, path, YANDEX_OAUTH_TOKEN, row)
        total_all += ch
        print(f"  Изменений: {ch}")

    # Итог
    ws.cell(row=row, column=1, value=f"Всего изменений: {total_all}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)

    # Автоширина по содержимому
    from openpyxl.cell.cell import MergedCell
    for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        max_len = 0
        col_letter = None
        for cell in col_cells:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 60)

    # Сохранить и загрузить
    out = io.BytesIO()
    wb_report.save(out)

    folder = os.path.dirname(DISK_SOURCE_PATH)
    ts_file = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%S")
    report_path = f"{folder}/diff_report_{ts_file}.xlsx"

    print(f"Upload report → {report_path}")
    disk_upload(report_path, out.getvalue(), YANDEX_OAUTH_TOKEN)
    print(f"Всего изменений: {total_all}")
    print("Done.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
