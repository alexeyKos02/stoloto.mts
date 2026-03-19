"""
Общие утилиты для синхронизации Excel-файлов через Яндекс Диск.
Используется всеми тремя скриптами: sync_inside_source, sync_full, sync_sheets.
"""

import io
import re
import time
from copy import copy
from typing import Dict, List, Optional, Set, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill


# =======================
# YANDEX DISK API
# =======================
YANDEX_API = "https://cloud-api.yandex.net/v1/disk"


def disk_download(path: str, token: str) -> bytes:
    headers = {"Authorization": f"OAuth {token}"}
    r = requests.get(
        f"{YANDEX_API}/resources/download",
        headers=headers,
        params={"path": path},
        timeout=60,
    )
    if r.status_code >= 400:
        raise RuntimeError(f"DOWNLOAD ERROR: {r.status_code}\nPATH: {path}\nBODY: {r.text}")
    href = r.json()["href"]

    f = requests.get(href, timeout=180)
    if f.status_code >= 400:
        raise RuntimeError(f"DOWNLOAD(HREF) ERROR: {f.status_code}\nHREF: {href}\nBODY: {f.text}")
    return f.content


def disk_upload(path: str, content: bytes, token: str, retries: int = 8) -> None:
    headers = {"Authorization": f"OAuth {token}"}
    r = requests.get(
        f"{YANDEX_API}/resources/upload",
        headers=headers,
        params={"path": path, "overwrite": "true"},
        timeout=60,
    )
    if r.status_code >= 400:
        raise RuntimeError(f"UPLOAD(HREF) ERROR: {r.status_code}\nPATH: {path}\nBODY: {r.text}")
    href = r.json()["href"]

    for attempt in range(1, retries + 1):
        put = requests.put(href, data=content, timeout=240)
        if put.status_code < 400:
            return

        if put.status_code == 423:
            wait = min(2 ** attempt, 30)
            print(f"  Upload LOCKED (423). Retry {attempt}/{retries} in {wait}s...")
            time.sleep(wait)
            continue

        raise RuntimeError(f"UPLOAD ERROR: {put.status_code}\nPATH: {path}\nBODY: {put.text}")

    raise RuntimeError(
        "UPLOAD ERROR: file is LOCKED too long (423). "
        "Закрой файл в Яндекс Таблицах/редакторе и запусти workflow ещё раз."
    )


# =======================
# HEADER / CELL HELPERS
# =======================
def header_index_map(ws: Worksheet) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            m[name] = c
    return m


def last_header_col(ws: Worksheet) -> int:
    """Последняя колонка в строке 1, где реально есть заголовок."""
    last = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip() != "":
            last = c
    return max(last, 1)


def is_empty_cell(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def get_cell_str(ws: Worksheet, r: int, c: int) -> str:
    v = ws.cell(row=r, column=c).value
    return "" if v is None else str(v).strip()


def get_last_data_row(ws: Worksheet, key_col: int, start_row: int = 2) -> int:
    """Последняя строка, где key_col заполнен (не max_row, который может быть огромным)."""
    last = 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if not is_empty_cell(v):
            last = r
    return last


def col_to_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


# =======================
# STYLE COPY
# =======================
def copy_cell_style(src_cell, dst_cell) -> None:
    """copy() чтобы не тащить StyleProxy и не словить unhashable StyleProxy на save()."""
    if not src_cell.has_style:
        return
    dst_cell._style = copy(src_cell._style)
    dst_cell.font = copy(src_cell.font)
    dst_cell.border = copy(src_cell.border)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.alignment = copy(src_cell.alignment)


def copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    """Копирует оформление строки (высота + стили ячеек)."""
    try:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    except Exception:
        pass
    for c in range(1, max_col + 1):
        copy_cell_style(ws.cell(row=src_row, column=c), ws.cell(row=dst_row, column=c))


def ensure_columns_at_end(ws: Worksheet, needed: List[str]) -> None:
    """Добавляем отсутствующие колонки в конец, копируя стиль/ширину с последней существующей."""
    m = header_index_map(ws)
    last = last_header_col(ws)

    template_col = last if last >= 1 else 1
    template_header = ws.cell(row=1, column=template_col)
    template_letter = col_to_letter(template_col)
    template_width = ws.column_dimensions[template_letter].width

    for name in needed:
        if name in m:
            continue
        last += 1
        dst_header = ws.cell(row=1, column=last)
        dst_header.value = name
        copy_cell_style(template_header, dst_header)
        new_letter = col_to_letter(last)
        if template_width is not None:
            ws.column_dimensions[new_letter].width = template_width
        m[name] = last


# =======================
# BOOL NORMALIZATION
# =======================
def normalize_bool_to_01(v) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return 1 if v else 0
    if isinstance(v, (int, float)):
        if v == 1:
            return 1
        if v == 0:
            return 0
    s = str(v).strip().lower()
    if s == "":
        return None
    if s in ("true", "истина", "да", "yes", "y", "1"):
        return 1
    if s in ("false", "ложь", "нет", "no", "n", "0"):
        return 0
    return None


# =======================
# CONDITIONAL FORMATTING (0/1)
# =======================
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GRAY = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")


def clear_col_cf(ws: Worksheet, col_letter: str) -> None:
    """Удаляет все CF-правила, в диапазоне которых участвует данная колонка."""
    # Матчим именно колонку col_letter, а не подстроку (E ≠ AE)
    pattern = re.compile(r'(?<![A-Z])' + re.escape(col_letter.upper()) + r'\d')
    to_remove = [
        key for key in list(ws.conditional_formatting._cf_rules.keys())
        if pattern.search(str(key).upper())
    ]
    for key in to_remove:
        del ws.conditional_formatting._cf_rules[key]


def apply_bool_cf(ws: Worksheet, col_letter: str, start_row: int, end_row: int) -> None:
    """CF: пусто→серый, 1→зелёный, 0→красный. Сначала чистит старые правила для этой колонки."""
    if end_row < start_row:
        end_row = start_row
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    r0 = start_row

    clear_col_cf(ws, col_letter)

    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'LEN(TRIM({col_letter}{r0}))=0'], fill=FILL_GRAY, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'{col_letter}{r0}=1'], fill=FILL_GREEN, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'{col_letter}{r0}=0'], fill=FILL_RED, stopIfTrue=False),
    )


# =======================
# TERMINAL ID RANGES
# =======================
def parse_terminal_id(x) -> Optional[int]:
    s = "".join(ch for ch in str(x) if ch.isdigit())
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None


def compress_ranges(nums: List[int]) -> List[Tuple[int, int]]:
    if not nums:
        return []
    nums = sorted(set(nums))
    out: List[Tuple[int, int]] = []
    start = prev = nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
            continue
        out.append((start, prev))
        start = prev = n
    out.append((start, prev))
    return out


def format_ranges(ranges: List[Tuple[int, int]]) -> str:
    parts = []
    for a, b in ranges:
        if a == b:
            parts.append(f"({a})")
        else:
            parts.append(f"({a}\u2013{b})")
    return " ".join(parts)


# =======================
# CERT FROM BD COMMENT
# =======================
CERT_OK_PHRASE = "есть все, но со стороны мтс нет сертификата"


def cert_value_from_bd_comment(comment_value) -> int:
    """
    Логика вычисления "Добавлен сертификат" по комментарию из БД:
    - пустой комментарий → 1
    - "есть все, но со стороны мтс нет сертификата" → 1
    - всё остальное → 0
    """
    s = "" if comment_value is None else str(comment_value).strip().lower()
    if s == "" or s == CERT_OK_PHRASE:
        return 1
    return 0


# =======================
# MTS ID NORMALIZATION
# =======================
def normalize_mts_id(value) -> str:
    """Ведущие нули: если число — добиваем до 9 цифр."""
    if value is None:
        return ""
    s = str(value).strip()
    if s == "":
        return ""
    digits = "".join(ch for ch in s if ch.isdigit())
    if digits == "":
        return ""
    if len(digits) > 9:
        return digits
    return digits.zfill(9)


# =======================
# TRANSLITERATION (ENG column in TARGET)
# =======================
def ru_to_translit(text: str) -> str:
    m = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "yo",
        "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
        "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
        "ф": "f", "х": "kh", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "shch",
        "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
        " ": "_", "_": "_", "-": "-",
    }
    s = (text or "").strip().lower()
    out = []
    for ch in s:
        if ch in m:
            out.append(m[ch])
        elif ch.isalnum() and (("a" <= ch <= "z") or ch.isdigit()):
            out.append(ch)
        else:
            out.append("_")
    slug = "".join(out)
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug.strip("_")
