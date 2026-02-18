import io
import os
import sys
import time
from typing import Dict, Optional, List

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill


# =======================
# ENV
# =======================
YANDEX_OAUTH_TOKEN = os.getenv("YANDEX_OAUTH_TOKEN", "").strip()
DISK_SOURCE_PATH = os.getenv("DISK_SOURCE_PATH", "").strip()
DISK_TARGET_PATH = os.getenv("DISK_TARGET_PATH", "").strip()

SRC_SHEET = os.getenv("SRC_SHEET", "СВОДНАЯ").strip()
TGT_SHEET = os.getenv("TGT_SHEET", "Лист1").strip()
KEY_COL = os.getenv("KEY_COL", "ЮЛ").strip()

COL_NAME = "Добавлен сертификат (МТС)"  # переносим ТОЛЬКО это

if not YANDEX_OAUTH_TOKEN:
    raise RuntimeError("ERROR: YANDEX_OAUTH_TOKEN is empty")
if not DISK_SOURCE_PATH:
    raise RuntimeError("ERROR: DISK_SOURCE_PATH is empty")
if not DISK_TARGET_PATH:
    raise RuntimeError("ERROR: DISK_TARGET_PATH is empty")

YANDEX_API = "https://cloud-api.yandex.net/v1/disk"
HEADERS = {"Authorization": f"OAuth {YANDEX_OAUTH_TOKEN}"}


# =======================
# Yandex Disk
# =======================
def disk_download(path: str) -> bytes:
    r = requests.get(
        f"{YANDEX_API}/resources/download",
        headers=HEADERS,
        params={"path": path},
        timeout=60,
    )
    if r.status_code >= 400:
        raise RuntimeError(f"DOWNLOAD ERROR {r.status_code}: {r.text}")
    href = r.json()["href"]

    f = requests.get(href, timeout=180)
    if f.status_code >= 400:
        raise RuntimeError(f"DOWNLOAD(HREF) ERROR {f.status_code}: {f.text}")
    return f.content


def disk_upload(path: str, content: bytes, retries: int = 8) -> None:
    r = requests.get(
        f"{YANDEX_API}/resources/upload",
        headers=HEADERS,
        params={"path": path, "overwrite": "true"},
        timeout=60,
    )
    if r.status_code >= 400:
        raise RuntimeError(f"UPLOAD(HREF) ERROR {r.status_code}: {r.text}")
    href = r.json()["href"]

    for attempt in range(1, retries + 1):
        put = requests.put(href, data=content, timeout=240)
        if put.status_code < 400:
            return
        if put.status_code == 423:
            wait = min(2 ** attempt, 30)
            print(f"⚠️ Upload LOCKED (423). Retry {attempt}/{retries} in {wait}s...")
            time.sleep(wait)
            continue
        raise RuntimeError(f"UPLOAD ERROR {put.status_code}: {put.text}")

    raise RuntimeError("UPLOAD ERROR: file LOCKED too long (423). Close it and rerun.")


# =======================
# Helpers
# =======================
def header_index_map(ws: Worksheet) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            m[name] = c
    return m


def ensure_column(ws: Worksheet, name: str) -> int:
    m = header_index_map(ws)
    if name in m:
        return m[name]
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = name
    return col


def is_empty(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def normalize_bool_to_01(v) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return 1 if v else 0
    if isinstance(v, (int, float)):
        if v == 1:
            return 1
        if v == 0:
            return 0
    s = str(v).strip().lower()
    if s == "":
        return None
    if s in ("true", "истина", "да", "yes", "y", "1"):
        return 1
    if s in ("false", "ложь", "нет", "no", "n", "0"):
        return 0
    return None


def get_last_data_row(ws: Worksheet, key_col: int, start_row: int = 2) -> int:
    last = 1
    for r in range(start_row, ws.max_row + 1):
        if not is_empty(ws.cell(row=r, column=key_col).value):
            last = r
    return last


def col_to_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


# =======================
# Conditional formatting (0/1)
# =======================
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GRAY = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")


def apply_bool_cf(ws: Worksheet, col_letter: str, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        end_row = start_row
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    r0 = start_row

    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'LEN(TRIM({col_letter}{r0}))=0'], fill=FILL_GRAY, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'{col_letter}{r0}=1'], fill=FILL_GREEN, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'{col_letter}{r0}=0'], fill=FILL_RED, stopIfTrue=False),
    )


# =======================
# Main sync
# =======================
def sync_target_to_source(source_bytes: bytes, target_bytes: bytes) -> bytes:
    wb_src = load_workbook(io.BytesIO(source_bytes))
    wb_tgt = load_workbook(io.BytesIO(target_bytes))

    if SRC_SHEET not in wb_src.sheetnames:
        raise RuntimeError(f'SOURCE: sheet "{SRC_SHEET}" not found')
    if TGT_SHEET not in wb_tgt.sheetnames:
        raise RuntimeError(f'TARGET: sheet "{TGT_SHEET}" not found')

    ws_src = wb_src[SRC_SHEET]
    ws_tgt = wb_tgt[TGT_SHEET]

    src_map = header_index_map(ws_src)
    tgt_map = header_index_map(ws_tgt)

    if KEY_COL not in src_map:
        raise RuntimeError(f'SOURCE: key column "{KEY_COL}" not found')
    if KEY_COL not in tgt_map:
        raise RuntimeError(f'TARGET: key column "{KEY_COL}" not found')

    src_key_c = src_map[KEY_COL]
    tgt_key_c = tgt_map[KEY_COL]

    # ensure col in both
    src_col_c = ensure_column(ws_src, COL_NAME)
    tgt_col_c = ensure_column(ws_tgt, COL_NAME)

    # build dict from TARGET
    tgt_last = get_last_data_row(ws_tgt, tgt_key_c, start_row=2)
    data: Dict[str, Optional[int]] = {}

    for r in range(2, tgt_last + 1):
        key = ws_tgt.cell(row=r, column=tgt_key_c).value
        if is_empty(key):
            continue
        v = ws_tgt.cell(row=r, column=tgt_col_c).value
        norm = normalize_bool_to_01(v)
        # если в target пусто/мусор — не тащим
        if norm is None:
            continue
        data[str(key).strip()] = norm

    # apply to SOURCE
    src_last = get_last_data_row(ws_src, src_key_c, start_row=2)
    updated = 0
    for r in range(2, src_last + 1):
        key = ws_src.cell(row=r, column=src_key_c).value
        if is_empty(key):
            continue
        k = str(key).strip()
        if k not in data:
            continue
        ws_src.cell(row=r, column=src_col_c).value = data[k]
        updated += 1

    # reapply CF in SOURCE only for this col
    letter = col_to_letter(src_col_c)
    apply_bool_cf(ws_src, letter, start_row=2, end_row=max(src_last, 2))

    print(f"TARGET -> SOURCE: updated={updated}, keys_with_value={len(data)}")

    out = io.BytesIO()
    wb_src.save(out)
    return out.getvalue()


def main() -> None:
    print(f"Download SOURCE: {DISK_SOURCE_PATH}")
    src = disk_download(DISK_SOURCE_PATH)
    print(f"Download TARGET: {DISK_TARGET_PATH}")
    tgt = disk_download(DISK_TARGET_PATH)

    out_src = sync_target_to_source(src, tgt)

    print(f"Upload SOURCE back: {DISK_SOURCE_PATH}")
    disk_upload(DISK_SOURCE_PATH, out_src)
    print("✅ Done")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        sys.exit(1)
