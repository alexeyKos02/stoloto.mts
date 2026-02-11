import os
import io
import requests
from openpyxl import load_workbook

YANDEX_TOKEN = os.environ["YANDEX_OAUTH_TOKEN"]
SOURCE_PATH = os.environ["DISK_SOURCE_PATH"]   # напр: /Apps/mts/source.xlsx
TARGET_PATH = os.environ["DISK_TARGET_PATH"]   # напр: /Apps/mts/target.xlsx

CONFIG = {
    "sourceSheet": "БД",
    "targetSheet": "СВОДНАЯ",
    "columns": ["ЮЛ", "МТС ID", "Terminal ID (Столото)", "Агент ID (Столото)", "GUID", "Ответственный ССПС"],
    "agentIdColumn": "Агент ID (Столото)",
    "dataStartRow": 2,
}

DISK_API = "https://cloud-api.yandex.net/v1/disk/resources"
HEADERS = {"Authorization": f"OAuth {YANDEX_TOKEN}"}


def normalize_mts_id(value) -> str:
    if value is None:
        return ""
    s = str(value)
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits or len(digits) > 9:
        return s
    return digits.zfill(9)


def disk_download(path: str) -> bytes:
    # Получаем ссылку на скачивание
    r = requests.get(f"{DISK_API}/download", headers=HEADERS, params={"path": path})
    r.raise_for_status()
    href = r.json()["href"]
    # Скачиваем файл
    f = requests.get(href)
    f.raise_for_status()
    return f.content


def disk_upload(path: str, content: bytes) -> None:
    # Получаем ссылку на загрузку
    r = requests.get(f"{DISK_API}/upload", headers=HEADERS, params={"path": path, "overwrite": "true"})
    r.raise_for_status()
    href = r.json()["href"]
    # Загружаем файл
    u = requests.put(href, data=content)
    u.raise_for_status()


def sheet_headers(ws, header_row=1):
    return [cell.value if cell.value is not None else "" for cell in ws[header_row]]


def get_col_indexes(headers, needed):
    idx = []
    for name in needed:
        if name not in headers:
            raise RuntimeError(f"Не найдена колонка: {name}")
        idx.append(headers.index(name))
    return idx


def row_is_empty(row_values):
    return all(v is None or str(v) == "" for v in row_values)


def main():
    # 1) Скачиваем оба файла
    src_bytes = disk_download(SOURCE_PATH)
    tgt_bytes = disk_download(TARGET_PATH)

    src_wb = load_workbook(io.BytesIO(src_bytes))
    tgt_wb = load_workbook(io.BytesIO(tgt_bytes))

    if CONFIG["sourceSheet"] not in src_wb.sheetnames:
        raise RuntimeError(f'Нет листа "{CONFIG["sourceSheet"]}" в источнике')
    if CONFIG["targetSheet"] not in tgt_wb.sheetnames:
        raise RuntimeError(f'Нет листа "{CONFIG["targetSheet"]}" в цели')

    src_ws = src_wb[CONFIG["sourceSheet"]]
    tgt_ws = tgt_wb[CONFIG["targetSheet"]]

    src_headers = sheet_headers(src_ws, 1)
    tgt_headers = sheet_headers(tgt_ws, 1)

    src_idx = get_col_indexes(src_headers, CONFIG["columns"])
    tgt_idx = get_col_indexes(tgt_headers, CONFIG["columns"])

    agent_src_i = src_headers.index(CONFIG["agentIdColumn"])
    agent_tgt_i = tgt_headers.index(CONFIG["agentIdColumn"])
    mts_pos = CONFIG["columns"].index("МТС ID")

    # 2) Собираем sourceMap: agentId -> rowData (уникальные)
    source_map = {}
    start = CONFIG["dataStartRow"]
    for r in range(start, src_ws.max_row + 1):
        row = [src_ws.cell(row=r, column=c+1).value for c in range(len(src_headers))]
        agent_id = row[agent_src_i]
        if agent_id is None or str(agent_id) == "":
            continue
        key = str(agent_id)
        if key in source_map:
            continue
        row_data = [row[i] for i in src_idx]
        row_data[mts_pos] = normalize_mts_id(row_data[mts_pos])
        source_map[key] = row_data

    # 3) Читаем target диапазон и решаем: update / clear / insert
    # Найдём все строки таргета (до max_row)
    updates = []      # (row_number, row_data)
    clears = []       # row_number
    existing_agent_rows = {}  # agentId -> row_number

    for r in range(start, tgt_ws.max_row + 1):
        row = [tgt_ws.cell(row=r, column=c+1).value for c in range(len(tgt_headers))]
        agent_id = row[agent_tgt_i]
        if agent_id is None or str(agent_id) == "":
            # пустая строка — кандидаты на вставку (но мы позже найдём первую пустую)
            continue
        existing_agent_rows[str(agent_id)] = r

    # Пройдём по заполненным строкам таргета и решим update/clear
    for agent_id, r in existing_agent_rows.items():
        if agent_id in source_map:
            src_row_data = source_map[agent_id]
            # сравнение по строке (как в GAS — через String)
            current = [tgt_ws.cell(row=r, column=i+1).value for i in tgt_idx]
            same = all(str(current[i]) == str(src_row_data[i]) for i in range(len(tgt_idx)))
            if not same:
                updates.append((r, src_row_data))
            del source_map[agent_id]
        else:
            clears.append(r)

    # 4) Очистки
    for r in clears:
        for col_i in tgt_idx:
            tgt_ws.cell(row=r, column=col_i+1).value = None

    # 5) Обновления
    for r, data in updates:
        for j, col_i in enumerate(tgt_idx):
            val = data[j]
            # МТС ID как текст: пишем строкой
            if CONFIG["columns"][j] == "МТС ID":
                val = normalize_mts_id(val)
            tgt_ws.cell(row=r, column=col_i+1).value = val

    # 6) Вставки в первую пустую строку и далее
    def find_first_empty():
        for r in range(start, tgt_ws.max_row + 2):
            row_vals = [tgt_ws.cell(row=r, column=i+1).value for i in tgt_idx]
            if row_is_empty(row_vals):
                return r
        return tgt_ws.max_row + 1

    insert_r = find_first_empty()
    inserted = 0
    for agent_id, data in source_map.items():
        for j, col_i in enumerate(tgt_idx):
            val = data[j]
            if CONFIG["columns"][j] == "МТС ID":
                val = normalize_mts_id(val)
            tgt_ws.cell(row=insert_r, column=col_i+1).value = val
        insert_r += 1
        inserted += 1

    # 7) Сохраняем и загружаем обратно
    out = io.BytesIO()
    tgt_wb.save(out)
    disk_upload(TARGET_PATH, out.getvalue())

    print(f"OK: cleared={len(clears)} updated={len(updates)} inserted={inserted}")


if __name__ == "__main__":
    main()
