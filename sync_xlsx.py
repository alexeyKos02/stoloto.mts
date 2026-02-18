import io
import os
import sys
import time
from copy import copy
from typing import Dict, List, Tuple, Optional, Set

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill


# =======================
# ENV (НЕ ПЕРЕИМЕНОВЫВАТЬ)
# =======================
YANDEX_OAUTH_TOKEN = os.getenv("YANDEX_OAUTH_TOKEN", "").strip()
DISK_SOURCE_PATH = os.getenv("DISK_SOURCE_PATH", "").strip()
DISK_TARGET_PATH = os.getenv("DISK_TARGET_PATH", "").strip()  # может не использоваться, но НЕ УДАЛЯТЬ

if not YANDEX_OAUTH_TOKEN:
    raise RuntimeError("ERROR: YANDEX_OAUTH_TOKEN is empty (set it in GitHub Secrets)")
if not DISK_SOURCE_PATH:
    raise RuntimeError("ERROR: DISK_SOURCE_PATH is empty (set it in GitHub Secrets)")

# Флаги запуска (можно управлять из workflow env)
RUN_INSIDE_SOURCE = os.getenv("RUN_INSIDE_SOURCE", "1").strip()  # 1 = обновляем СВОДНАЯ внутри SOURCE
RUN_SYNC_TO_TARGET = os.getenv("RUN_SYNC_TO_TARGET", "0").strip()  # 1 = синк SOURCE->TARGET (отдельно)

# Настройки синка SOURCE->TARGET (если включишь RUN_SYNC_TO_TARGET=1)
SRC_SHEET_FOR_EXPORT = os.getenv("SRC_SHEET_FOR_EXPORT", "СВОДНАЯ").strip()
TGT_SHEET_FOR_IMPORT = os.getenv("TGT_SHEET_FOR_IMPORT", "Лист1").strip()
KEY_COLUMN_EXPORT = os.getenv("KEY_COLUMN_EXPORT", "ЮЛ").strip()
COLUMNS_TO_SYNC_EXPORT = os.getenv(
    "COLUMNS_TO_SYNC_EXPORT",
    "ЮЛ|Terminal ID (Столото)|МТС ID|Добавлен сертификат|Добавлен сертификат (МТС)|Билеты продаются",
).strip()


YANDEX_API = "https://cloud-api.yandex.net/v1/disk"
HEADERS = {"Authorization": f"OAuth {YANDEX_OAUTH_TOKEN}"}


# =======================
# CONFIG (ЛИСТЫ/КОЛОНКИ)
# =======================
SHEET_BD = "БД"
SHEET_SVOD = "СВОДНАЯ"

SVOD_BOOL_COLS = [
    "Добавлен сертификат",
    "Добавлен сертификат (МТС)",
    "Билеты продаются",
]

SVOD_REQUIRED_BASE = [
    "ЮЛ",
    "МТС ID",
    "Terminal ID (Столото)",
    "Агент ID (Столото)",
    "GUID",
    "Ответственный ССПС",
]

BD_REQUIRED = [
    "ЮЛ",
    "МТС ID",
    "Terminal ID (Столото)",
    "Агент ID (Столото)",
    "GUID",
    "Ответственный ССПС",
]


# =======================
# YANDEX DISK API
# =======================
def disk_download(path: str) -> bytes:
    r = requests.get(
        f"{YANDEX_API}/resources/download",
        headers=HEADERS,
        params={"path": path},
        timeout=60,
    )
    if r.status_code >= 400:
        raise RuntimeError(f"DOWNLOAD ERROR: {r.status_code}\nPATH: {path}\nBODY: {r.text}")
    href = r.json()["href"]

    f = requests.get(href, timeout=180)
    if f.status_code >= 400:
        raise RuntimeError(f"DOWNLOAD(HREF) ERROR: {f.status_code}\nHREF: {href}\nBODY: {f.text}")
    return f.content


def disk_upload(path: str, content: bytes, retries: int = 8) -> None:
    r = requests.get(
        f"{YANDEX_API}/resources/upload",
        headers=HEADERS,
        params={"path": path, "overwrite": "true"},
        timeout=60,
    )
    if r.status_code >= 400:
        raise RuntimeError(f"UPLOAD(HREF) ERROR: {r.status_code}\nPATH: {path}\nBODY: {r.text}")
    href = r.json()["href"]

    for attempt in range(1, retries + 1):
        put = requests.put(href, data=content, timeout=240)
        if put.status_code < 400:
            return

        if put.status_code == 423:
            wait = min(2**attempt, 30)
            print(f"⚠️ Upload LOCKED (423). Retry {attempt}/{retries} in {wait}s...")
            time.sleep(wait)
            continue

        raise RuntimeError(f"UPLOAD ERROR: {put.status_code}\nPATH: {path}\nBODY: {put.text}")

    raise RuntimeError(
        "UPLOAD ERROR: file is LOCKED too long (423). "
        "Закрой файл в Яндекс Таблицах/редакторе и запусти workflow ещё раз."
    )


# =======================
# HELPERS: columns / values
# =======================
def header_index_map(ws: Worksheet) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            m[name] = c
    return m


def is_empty_cell(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def get_cell_str(ws: Worksheet, r: int, c: int) -> str:
    v = ws.cell(row=r, column=c).value
    return "" if v is None else str(v).strip()


def col_to_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def ru_to_translit(text: str) -> str:
    m = {
        "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "yo",
        "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
        "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
        "ф": "f", "х": "kh", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "shch",
        "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
        " ": "_", "_": "_", "-": "-"
    }

    s = (text or "").strip().lower()
    out = []
    for ch in s:
        if ch in m:
            out.append(m[ch])
        elif ch.isalnum() and (("a" <= ch <= "z") or ch.isdigit()):
            out.append(ch)
        else:
            out.append("_")

    slug = "".join(out)
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug.strip("_")

def last_header_col(ws: Worksheet) -> int:
    """Последняя колонка в строке 1, где реально есть заголовок (value)."""
    last = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip() != "":
            last = c
    return last


# =======================
# STYLE COPY (FIX StyleProxy crash)
# =======================
def copy_cell_style(src_cell, dst_cell) -> None:
    """
    Важно: copy() чтобы не тащить StyleProxy и не словить unhashable StyleProxy на save().
    """
    if not src_cell.has_style:
        return

    dst_cell._style = copy(src_cell._style)

    # продублируем явно (это нормально и безопасно)
    dst_cell.font = copy(src_cell.font)
    dst_cell.border = copy(src_cell.border)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.alignment = copy(src_cell.alignment)


def copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    """
    Копирует оформление строки (высота + стили ячеек) из src_row в dst_row.
    """
    # высота строки
    try:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    except Exception:
        pass

    for c in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=c)
        d = ws.cell(row=dst_row, column=c)
        copy_cell_style(s, d)


def ensure_columns_at_end(ws: Worksheet, needed: List[str]) -> None:
    """
    Добавляем отсутствующие колонки в конец.
    Чтобы не ломать форматирование/ширины — копируем стиль заголовка и ширину
    с последней существующей колонки.
    """
    m = header_index_map(ws)
    last = ws.max_column

    # шаблонный заголовок и ширина — от последней существующей колонки
    template_col = last if last >= 1 else 1
    template_header = ws.cell(row=1, column=template_col)
    template_letter = col_to_letter(template_col)
    template_width = ws.column_dimensions[template_letter].width

    for name in needed:
        if name in m:
            continue
        last += 1

        dst_header = ws.cell(row=1, column=last)
        dst_header.value = name

        # стиль заголовка
        copy_cell_style(template_header, dst_header)

        # ширина
        new_letter = col_to_letter(last)
        if template_width is not None:
            ws.column_dimensions[new_letter].width = template_width

        m[name] = last


# =======================
# FIX: "последняя строка данных", а не max_row
# =======================
def get_last_data_row(ws: Worksheet, key_col: int, start_row: int = 2) -> int:
    last = 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if not is_empty_cell(v):
            last = r
    return last


# =======================
# TERMINAL RANGES
# =======================
def parse_terminal_id(x: str) -> Optional[int]:
    s = "".join(ch for ch in str(x) if ch.isdigit())
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None


def compress_ranges(nums: List[int]) -> List[Tuple[int, int]]:
    if not nums:
        return []
    nums = sorted(set(nums))
    out: List[Tuple[int, int]] = []
    start = prev = nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
            continue
        out.append((start, prev))
        start = prev = n
    out.append((start, prev))
    return out


def format_ranges(ranges: List[Tuple[int, int]]) -> str:
    parts = []
    for a, b in ranges:
        if a == b:
            parts.append(f"({a})")
        else:
            parts.append(f"({a}–{b})")
    return " ".join(parts)


# =======================
# CONDITIONAL FORMATTING (0/1)
# =======================
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GRAY = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")


def apply_bool_cf(ws: Worksheet, col_letter: str, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        end_row = start_row
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    r0 = start_row

    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'LEN(TRIM({col_letter}{r0}))=0'], fill=FILL_GRAY, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'{col_letter}{r0}=1'], fill=FILL_GREEN, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'{col_letter}{r0}=0'], fill=FILL_RED, stopIfTrue=False),
    )


def normalize_bool_to_01(v) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return 1 if v else 0
    if isinstance(v, (int, float)):
        if v == 1:
            return 1
        if v == 0:
            return 0
    s = str(v).strip().lower()
    if s == "":
        return None
    if s in ("true", "истина", "да", "yes", "y", "1"):
        return 1
    if s in ("false", "ложь", "нет", "no", "n", "0"):
        return 0
    return None


def ensure_svod_columns(ws_svod: Worksheet) -> None:
    ensure_columns_at_end(ws_svod, SVOD_BOOL_COLS)


# =======================
# DELETE agents removed from BD
# =======================
def delete_missing_agents(ws_svod: Worksheet, sv_map: Dict[str, int], agents_in_bd: Set[str]) -> int:
    agent_col = sv_map["Агент ID (Столото)"]
    last_data = get_last_data_row(ws_svod, agent_col, start_row=2)
    if last_data < 2:
        return 0

    to_delete: List[int] = []
    for r in range(2, last_data + 1):
        agent = get_cell_str(ws_svod, r, agent_col)
        if agent and agent not in agents_in_bd:
            to_delete.append(r)

    deleted = 0
    for r in reversed(to_delete):
        ws_svod.delete_rows(r, 1)
        deleted += 1

    return deleted


# =======================
# MAIN SYNC LOGIC (inside SOURCE)
# =======================
def sync_inside_workbook(src_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(src_bytes))

    if SHEET_BD not in wb.sheetnames:
        raise RuntimeError(f'Source: sheet "{SHEET_BD}" not found')
    if SHEET_SVOD not in wb.sheetnames:
        raise RuntimeError(f'Target: sheet "{SHEET_SVOD}" not found')

    ws_bd = wb[SHEET_BD]
    ws_svod = wb[SHEET_SVOD]

    print(f'Ensure columns in "{SHEET_SVOD}"...')
    ensure_svod_columns(ws_svod)

    bd_map = header_index_map(ws_bd)
    sv_map = header_index_map(ws_svod)

    missing_bd = [c for c in BD_REQUIRED if c not in bd_map]
    if missing_bd:
        raise RuntimeError(f'Missing columns in "{SHEET_BD}": {missing_bd}')

    missing_svod = [c for c in SVOD_REQUIRED_BASE if c not in sv_map]
    if missing_svod:
        raise RuntimeError(f'Missing columns in "{SHEET_SVOD}": {missing_svod}')

    agent_col_bd = bd_map["Агент ID (Столото)"]
    terminal_col_bd = bd_map["Terminal ID (Столото)"]

    bd_by_agent: Dict[str, Dict[str, str]] = {}
    terminals_by_agent: Dict[str, List[int]] = {}
    agents_in_bd: Set[str] = set()

    for r in range(2, ws_bd.max_row + 1):
        agent = get_cell_str(ws_bd, r, agent_col_bd)
        if not agent:
            continue

        agents_in_bd.add(agent)

        term_raw = ws_bd.cell(row=r, column=terminal_col_bd).value
        term_num = parse_terminal_id(term_raw) if term_raw is not None else None
        if term_num is not None:
            terminals_by_agent.setdefault(agent, []).append(term_num)

        payload = bd_by_agent.setdefault(agent, {k: "" for k in BD_REQUIRED})
        for col_name in BD_REQUIRED:
            val = get_cell_str(ws_bd, r, bd_map[col_name])
            if payload[col_name] == "" and val != "":
                payload[col_name] = val

    for agent, nums in terminals_by_agent.items():
        rngs = compress_ranges(nums)
        bd_by_agent[agent]["Terminal ID (Столото)"] = format_ranges(rngs)

    deleted = delete_missing_agents(ws_svod, sv_map, agents_in_bd)
    if deleted:
        print(f"Deleted from SVOD (not in BD): {deleted}")

    sv_map = header_index_map(ws_svod)
    agent_col_sv = sv_map["Агент ID (Столото)"]

    last_data_row = get_last_data_row(ws_svod, agent_col_sv, start_row=2)

    # Шаблон строки ДАННЫХ — строго строка 2 (если она существует), иначе last_data_row
    if ws_svod.max_row >= 2:
        template_row = 2
    else:
        template_row = last_data_row if last_data_row >= 2 else 2

    max_col = ws_svod.max_column

    existing_row_by_agent: Dict[str, int] = {}
    if last_data_row >= 2:
        for r in range(2, last_data_row + 1):
            agent = get_cell_str(ws_svod, r, agent_col_sv)
            if agent:
                existing_row_by_agent[agent] = r

    inserted = 0
    updated = 0

    append_row = last_data_row + 1 if last_data_row >= 2 else 2

    for agent, payload in bd_by_agent.items():
        if agent in existing_row_by_agent:
            rr = existing_row_by_agent[agent]
            for col_name in SVOD_REQUIRED_BASE:
                ws_svod.cell(row=rr, column=sv_map[col_name]).value = payload.get(col_name, "")
            updated += 1
        else:
            rr = append_row
            append_row += 1

            # копируем оформление строки-образца (чтобы сетка/заливка/высота не ломались)
            if template_row >= 2 and template_row <= ws_svod.max_row:
                copy_row_style(ws_svod, template_row, rr, max_col)

            for col_name in SVOD_REQUIRED_BASE:
                ws_svod.cell(row=rr, column=sv_map[col_name]).value = payload.get(col_name, "")
            # новые 3 столбца: по умолчанию = 0
            for col_name in SVOD_BOOL_COLS:
                ws_svod.cell(row=rr, column=sv_map[col_name]).value = 0

            inserted += 1

    # нормализация 0/1 только по реальным данным
    last_data_row = get_last_data_row(ws_svod, agent_col_sv, start_row=2)
    for col_name in SVOD_BOOL_COLS:
        c = sv_map[col_name]
        for r in range(2, last_data_row + 1):
            v = ws_svod.cell(row=r, column=c).value
            if is_empty_cell(v):
                continue
            norm = normalize_bool_to_01(v)
            if norm is None:
                continue
            ws_svod.cell(row=r, column=c).value = norm

    # CF на реальные строки данных
    for col_name in SVOD_BOOL_COLS:
        c = sv_map[col_name]
        letter = col_to_letter(c)
        apply_bool_cf(ws_svod, letter, start_row=2, end_row=max(last_data_row, 2))

    print(
        f"Inside sync done: inserted={inserted}, updated={updated}, deleted={deleted}, "
        f"total_source_agents={len(bd_by_agent)}"
    )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =======================
# SYNC SOURCE -> TARGET (2nd file)
# =======================
def parse_columns_list(s: str) -> List[str]:
    parts = [p.strip() for p in s.split("|")]
    return [p for p in parts if p]


def sync_source_to_target(source_bytes: bytes, target_bytes: bytes) -> bytes:
    wb_src = load_workbook(io.BytesIO(source_bytes))
    wb_tgt = load_workbook(io.BytesIO(target_bytes))

    if SRC_SHEET_FOR_EXPORT not in wb_src.sheetnames:
        raise RuntimeError(f'Source file: sheet "{SRC_SHEET_FOR_EXPORT}" not found')
    ws_src = wb_src[SRC_SHEET_FOR_EXPORT]

    ws_tgt = (
        wb_tgt[TGT_SHEET_FOR_IMPORT]
        if TGT_SHEET_FOR_IMPORT in wb_tgt.sheetnames
        else wb_tgt.create_sheet(TGT_SHEET_FOR_IMPORT)
    )

    # --- настройки колонок ---
    ENG_COL = "ENG"
    UL_COL = "ЮЛ"
    BOOL_COLS = ["Добавлен сертификат", "Добавлен сертификат (МТС)", "Билеты продаются"]

    cols_base = parse_columns_list(COLUMNS_TO_SYNC_EXPORT)

    # ключ обязан быть в списке синкаемых
    if KEY_COLUMN_EXPORT not in cols_base:
        cols_base = [KEY_COLUMN_EXPORT] + cols_base

    # cols = только то, что реально синкаем из SOURCE (без ENG)
    cols = cols_base.copy()

    # Булевые переносим в TARGET и синкаем (если есть в SOURCE), иначе будет 0
    for b in BOOL_COLS:
        if b not in cols:
            cols.append(b)

    src_map = header_index_map(ws_src)
    tgt_map = header_index_map(ws_tgt)

    if KEY_COLUMN_EXPORT not in src_map:
        raise RuntimeError(
            f'Source sheet "{SRC_SHEET_FOR_EXPORT}": key column "{KEY_COLUMN_EXPORT}" not found'
        )

    # --- ensure headers in TARGET: (cols + ENG), добавляем рядом с последним реальным заголовком ---
    tgt_map = header_index_map(ws_tgt)
    h_last = last_header_col(ws_tgt)

    def ensure_header(name: str) -> None:
        nonlocal h_last, tgt_map
        if name in tgt_map:
            return
        h_last += 1
        ws_tgt.cell(row=1, column=h_last).value = name
        tgt_map[name] = h_last  # локально обновим

    # 1) сначала базовые колонки (без булевых)
    for name in cols_base:
        ensure_header(name)

    # 2) потом ENG (только в TARGET)
    ensure_header(ENG_COL)

    # 3) потом 3 булевых
    for b in BOOL_COLS:
        ensure_header(b)


    # refresh maps after header changes
    tgt_map = header_index_map(ws_tgt)

    # --- границы данных ---
    src_last = get_last_data_row(ws_src, src_map[KEY_COLUMN_EXPORT], start_row=2)
    tgt_last = get_last_data_row(ws_tgt, tgt_map[KEY_COLUMN_EXPORT], start_row=2)

    # --- читаем SOURCE в dict по ключу ---
    src_data: Dict[str, Dict[str, str]] = {}
    for r in range(2, src_last + 1):
        key = get_cell_str(ws_src, r, src_map[KEY_COLUMN_EXPORT])
        if not key:
            continue

        row_payload: Dict[str, str] = {}
        for col in cols:
            # если колонки нет в SOURCE — пишем пусто (не упадём)
            row_payload[col] = get_cell_str(ws_src, r, src_map[col]) if col in src_map else ""
        src_data[key] = row_payload

    # --- существующие строки TARGET по ключу ---
    tgt_row_by_key: Dict[str, int] = {}
    if tgt_last >= 2:
        for r in range(2, tgt_last + 1):
            key = get_cell_str(ws_tgt, r, tgt_map[KEY_COLUMN_EXPORT])
            if key:
                tgt_row_by_key[key] = r

    # --- шаблон для стилей ---
    template_row = 2 if ws_tgt.max_row >= 2 else 2
    max_col = last_header_col(ws_tgt)  # важно: по реальным заголовкам, а не max_column

    updated = 0
    inserted = 0
    append_row = tgt_last + 1 if tgt_last >= 2 else 2

    # --- upsert ---
    for key, payload in src_data.items():
        if key in tgt_row_by_key:
            rr = tgt_row_by_key[key]
            # обновляем только синкаемые cols (ENG не трогаем)
            for col in cols:
                ws_tgt.cell(row=rr, column=tgt_map[col]).value = payload.get(col, "")
            updated += 1
        else:
            rr = append_row
            append_row += 1

            # стиль строки
            if template_row >= 2 and template_row <= ws_tgt.max_row:
                copy_row_style(ws_tgt, template_row, rr, max_col)

            # записываем синкаемые колонки
            for col in cols:
                val = payload.get(col, "")

                # для булевых по умолчанию ставим 0, если из SOURCE пусто
                if col in BOOL_COLS and (val is None or str(val).strip() == ""):
                    val = 0

                ws_tgt.cell(row=rr, column=tgt_map[col]).value = val

            inserted += 1

    # --- normalize BOOL_COLS to 0/1 in TARGET + default 0 for empty ---
    key_col = tgt_map[KEY_COLUMN_EXPORT]
    tgt_last = get_last_data_row(ws_tgt, key_col, start_row=2)
    tgt_last = max(tgt_last, 2)

    for name in BOOL_COLS:
        if name not in tgt_map:
            continue
        c = tgt_map[name]
        for r in range(2, tgt_last + 1):
            v = ws_tgt.cell(row=r, column=c).value

            # пусто -> ставим 0 (твой запрос “по умолчанию 0”)
            if is_empty_cell(v):
                ws_tgt.cell(row=r, column=c).value = 0
                continue

            norm = normalize_bool_to_01(v)
            if norm is None:
                continue
            ws_tgt.cell(row=r, column=c).value = norm

    # --- re-apply conditional formatting in TARGET ---
    for name in BOOL_COLS:
        if name not in tgt_map:
            continue
        c = tgt_map[name]
        letter = col_to_letter(c)
        apply_bool_cf(ws_tgt, letter, start_row=2, end_row=tgt_last)

    # --- AUTOTRANSLIT ONLY IN TARGET: fill ENG if empty ---
    if UL_COL in tgt_map and ENG_COL in tgt_map:
        ul_c = tgt_map[UL_COL]
        eng_c = tgt_map[ENG_COL]

        last_ul = get_last_data_row(ws_tgt, ul_c, start_row=2)

        filled = 0
        for r in range(2, last_ul + 1):
            ul_val = ws_tgt.cell(row=r, column=ul_c).value
            eng_val = ws_tgt.cell(row=r, column=eng_c).value

            if is_empty_cell(ul_val):
                continue
            if not is_empty_cell(eng_val):
                continue

            ws_tgt.cell(row=r, column=eng_c).value = ru_to_translit(str(ul_val))
            filled += 1

        print(f"ENG filled: {filled}")
    else:
        print("ENG/ЮЛ columns not found in TARGET — skip ENG fill")

    print(f"SOURCE->TARGET sync done: updated={updated}, inserted={inserted}, total_source={len(src_data)}")

    out = io.BytesIO()
    wb_tgt.save(out)
    return out.getvalue()



# =======================
# ENTRYPOINT
# =======================
def main() -> None:
    inside = RUN_INSIDE_SOURCE == "1"
    to_target = RUN_SYNC_TO_TARGET == "1"

    if inside:
        print(f"Download SOURCE: {DISK_SOURCE_PATH}")
        src = disk_download(DISK_SOURCE_PATH)
        print(f"downloaded SOURCE: {len(src)} bytes")

        print("Running inside SOURCE sync...")
        out = sync_inside_workbook(src)

        print(f"Upload back to same path (SOURCE): {DISK_SOURCE_PATH}")
        disk_upload(DISK_SOURCE_PATH, out)
        print("✅ Inside SOURCE done")

    if to_target:
        if not DISK_TARGET_PATH:
            raise RuntimeError("ERROR: DISK_TARGET_PATH is empty (set it in GitHub Secrets)")

        print(f"Download SOURCE: {DISK_SOURCE_PATH}")
        src = disk_download(DISK_SOURCE_PATH)
        print(f"downloaded SOURCE: {len(src)} bytes")

        print(f"Download TARGET: {DISK_TARGET_PATH}")
        tgt = disk_download(DISK_TARGET_PATH)
        print(f"downloaded TARGET: {len(tgt)} bytes")

        print("Running SOURCE -> TARGET sync...")
        out_tgt = sync_source_to_target(src, tgt)

        print(f"Upload TARGET back: {DISK_TARGET_PATH}")
        disk_upload(DISK_TARGET_PATH, out_tgt)
        print("✅ SOURCE->TARGET done")

    if not inside and not to_target:
        print("⚠️ Nothing to do: set RUN_INSIDE_SOURCE=1 and/or RUN_SYNC_TO_TARGET=1")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        sys.exit(1)
