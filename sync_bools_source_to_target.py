import io
import os
import sys
import time
from typing import Dict, Optional, List, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill


# =======================
# ENV (НЕ ПЕРЕИМЕНОВЫВАТЬ)
# =======================
YANDEX_OAUTH_TOKEN = os.getenv("YANDEX_OAUTH_TOKEN", "").strip()
DISK_SOURCE_PATH = os.getenv("DISK_SOURCE_PATH", "").strip()
DISK_TARGET_PATH = os.getenv("DISK_TARGET_PATH", "").strip()

SRC_SHEET = os.getenv("SRC_SHEET", "СВОДНАЯ").strip()
TGT_SHEET = os.getenv("TGT_SHEET", "Лист1").strip()
KEY_COL = os.getenv("KEY_COL", "ЮЛ").strip()

# Какие колонки ПЕРЕНОСИМ значениями из SOURCE -> TARGET
COLS_SYNC = ["Добавлен сертификат", "Билеты продаются"]

# На каких колонках ДОЛЖНО ОСТАВАТЬСЯ цветное условное форматирование в TARGET
# (Добавлен сертификат (МТС) не трогаем значениями, но форматирование восстанавливаем)
COLS_WITH_CF = ["Добавлен сертификат", "Добавлен сертификат (МТС)", "Билеты продаются"]

if not YANDEX_OAUTH_TOKEN:
    raise RuntimeError("ERROR: YANDEX_OAUTH_TOKEN is empty")
if not DISK_SOURCE_PATH:
    raise RuntimeError("ERROR: DISK_SOURCE_PATH is empty")
if not DISK_TARGET_PATH:
    raise RuntimeError("ERROR: DISK_TARGET_PATH is empty")

YANDEX_API = "https://cloud-api.yandex.net/v1/disk"
HEADERS = {"Authorization": f"OAuth {YANDEX_OAUTH_TOKEN}"}


# =======================
# YANDEX DISK API
# =======================
def disk_download(path: str) -> bytes:
    r = requests.get(f"{YANDEX_API}/resources/download", headers=HEADERS, params={"path": path}, timeout=60)
    if r.status_code >= 400:
        raise RuntimeError(f"DOWNLOAD ERROR {r.status_code}: {r.text}")
    href = r.json()["href"]
    f = requests.get(href, timeout=180)
    if f.status_code >= 400:
        raise RuntimeError(f"DOWNLOAD(HREF) ERROR {f.status_code}: {f.text}")
    return f.content


def disk_upload(path: str, content: bytes, retries: int = 8) -> None:
    r = requests.get(
        f"{YANDEX_API}/resources/upload",
        headers=HEADERS,
        params={"path": path, "overwrite": "true"},
        timeout=60,
    )
    if r.status_code >= 400:
        raise RuntimeError(f"UPLOAD(HREF) ERROR {r.status_code}: {r.text}")
    href = r.json()["href"]

    for attempt in range(1, retries + 1):
        put = requests.put(href, data=content, timeout=240)
        if put.status_code < 400:
            return
        if put.status_code == 423:
            wait = min(2 ** attempt, 30)
            print(f"⚠️ Upload LOCKED (423). Retry {attempt}/{retries} in {wait}s...")
            time.sleep(wait)
            continue
        raise RuntimeError(f"UPLOAD ERROR {put.status_code}: {put.text}")

    raise RuntimeError("UPLOAD ERROR: file LOCKED too long (423). Close it and rerun.")


# =======================
# HELPERS: sheet / values
# =======================
def header_index_map(ws: Worksheet) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            m[name] = c
    return m


def last_header_col(ws: Worksheet) -> int:
    """
    Последний НЕ пустой заголовок в первой строке.
    Это важно, чтобы не уехать в 'пустые колонки', которые тянутся из-за форматирования.
    """
    last = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        if str(v).strip() != "":
            last = c
    return last if last > 0 else 1


def ensure_column(ws: Worksheet, name: str) -> int:
    """
    Гарантирует колонку в TARGET. Добавляет В КОНЕЦ реальных заголовков.
    """
    m = header_index_map(ws)
    if name in m:
        return m[name]
    col = last_header_col(ws) + 1
    ws.cell(row=1, column=col).value = name
    return col


def get_cell_str(ws: Worksheet, r: int, c: int) -> str:
    v = ws.cell(row=r, column=c).value
    return "" if v is None else str(v).strip()


def is_empty_cell(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def get_last_data_row(ws: Worksheet, key_col: int, start_row: int = 2) -> int:
    """
    Последняя строка, где key_col заполнен.
    Лечит 'max_row огромный из-за форматирования'.
    """
    last = 1
    for r in range(start_row, ws.max_row + 1):
        if not is_empty_cell(ws.cell(row=r, column=key_col).value):
            last = r
    return last


# =======================
# BOOL normalization + CF
# =======================
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GRAY = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")


def normalize_bool_to_01(v) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return 1 if v else 0
    if isinstance(v, (int, float)):
        if v == 1:
            return 1
        if v == 0:
            return 0
    s = str(v).strip().lower()
    if s == "":
        return None
    if s in ("true", "истина", "да", "yes", "y", "1"):
        return 1
    if s in ("false", "ложь", "нет", "no", "n", "0"):
        return 0
    return None


def col_to_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _remove_cf_rules_for_ranges(ws: Worksheet, ranges: List[str]) -> None:
    """
    Убираем существующие CF правила на указанные ranges, чтобы не копить дубликаты.
    openpyxl официально не даёт нормального удаления по диапазону, поэтому используем внутренности
    максимально аккуратно.
    """
    try:
        cf = ws.conditional_formatting
        rules = getattr(cf, "_cf_rules", None)
        if not isinstance(rules, dict):
            return
        for rng in ranges:
            if rng in rules:
                del rules[rng]
    except Exception:
        # Если внутренности изменились — просто не удаляем (хуже не станет, кроме дубликатов)
        pass


def apply_bool_cf(ws: Worksheet, col_letter: str, start_row: int, end_row: int) -> None:
    """
    CF:
      пусто -> серый
      1 -> зелёный
      0 -> красный
    Важно: формулы должны ссылаться на первую строку диапазона (Excel/Яндекс протянут).
    """
    if end_row < start_row:
        end_row = start_row
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    r0 = start_row

    _remove_cf_rules_for_ranges(ws, [rng])

    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'LEN(TRIM({col_letter}{r0}))=0'], fill=FILL_GRAY, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'{col_letter}{r0}=1'], fill=FILL_GREEN, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'{col_letter}{r0}=0'], fill=FILL_RED, stopIfTrue=False),
    )


# =======================
# SYNC
# =======================
def sync(source_bytes: bytes, target_bytes: bytes) -> bytes:
    wb_src = load_workbook(io.BytesIO(source_bytes))
    wb_tgt = load_workbook(io.BytesIO(target_bytes))

    if SRC_SHEET not in wb_src.sheetnames:
        raise RuntimeError(f'Source file: sheet "{SRC_SHEET}" not found')
    if TGT_SHEET not in wb_tgt.sheetnames:
        raise RuntimeError(f'Target file: sheet "{TGT_SHEET}" not found')

    ws_src = wb_src[SRC_SHEET]
    ws_tgt = wb_tgt[TGT_SHEET]

    src_map = header_index_map(ws_src)
    tgt_map = header_index_map(ws_tgt)

    if KEY_COL not in src_map:
        raise RuntimeError(f'Source sheet "{SRC_SHEET}": key column "{KEY_COL}" not found')

    # 1) гарантируем колонки в TARGET
    ensure_column(ws_tgt, KEY_COL)
    for c in COLS_SYNC:
        ensure_column(ws_tgt, c)
    for c in COLS_WITH_CF:
        ensure_column(ws_tgt, c)

    # refresh maps after headers
    src_map = header_index_map(ws_src)
    tgt_map = header_index_map(ws_tgt)

    # 2) границы данных
    src_last = get_last_data_row(ws_src, src_map[KEY_COL], start_row=2)
    tgt_last = get_last_data_row(ws_tgt, tgt_map[KEY_COL], start_row=2)

    # 3) source map: key -> values (только COLS_SYNC)
    src_data: Dict[str, Dict[str, Optional[int]]] = {}
    for r in range(2, src_last + 1):
        key = get_cell_str(ws_src, r, src_map[KEY_COL])
        if not key:
            continue
        payload: Dict[str, Optional[int]] = {}
        for name in COLS_SYNC:
            if name not in src_map:
                payload[name] = None
                continue
            payload[name] = normalize_bool_to_01(ws_src.cell(row=r, column=src_map[name]).value)
        src_data[key] = payload

    # 4) existing target rows by key
    tgt_row_by_key: Dict[str, int] = {}
    if tgt_last >= 2:
        for r in range(2, tgt_last + 1):
            key = get_cell_str(ws_tgt, r, tgt_map[KEY_COL])
            if key:
                tgt_row_by_key[key] = r

    # 5) upsert
    updated = 0
    inserted = 0
    append_row = tgt_last + 1 if tgt_last >= 2 else 2

    for key, payload in src_data.items():
        if key in tgt_row_by_key:
            rr = tgt_row_by_key[key]
            for name in COLS_SYNC:
                # переносим 0/1, пустое оставляем пустым
                val = payload.get(name, None)
                if val is None:
                    continue
                ws_tgt.cell(row=rr, column=tgt_map[name]).value = val
            updated += 1
        else:
            rr = append_row
            append_row += 1
            ws_tgt.cell(row=rr, column=tgt_map[KEY_COL]).value = key
            for name in COLS_SYNC:
                val = payload.get(name, None)
                ws_tgt.cell(row=rr, column=tgt_map[name]).value = (val if val is not None else None)
            inserted += 1

    # 6) нормализуем уже существующие значения в колонках CF (включая "(МТС)") к 0/1
    # НО: не перезаписываем пустые
    new_tgt_last = get_last_data_row(ws_tgt, tgt_map[KEY_COL], start_row=2)
    new_tgt_last = max(new_tgt_last, 2)

    for name in COLS_WITH_CF:
        if name not in tgt_map:
            continue
        c = tgt_map[name]
        for r in range(2, new_tgt_last + 1):
            v = ws_tgt.cell(row=r, column=c).value
            if is_empty_cell(v):
                continue
            norm = normalize_bool_to_01(v)
            if norm is None:
                continue
            ws_tgt.cell(row=r, column=c).value = norm

    # 7) восстанавливаем условное форматирование на ВСЕ три колонки (включая "(МТС)")
    for name in COLS_WITH_CF:
        if name not in tgt_map:
            continue
        letter = col_to_letter(tgt_map[name])
        apply_bool_cf(ws_tgt, letter, start_row=2, end_row=new_tgt_last)

    print(f"Sync done: updated={updated}, inserted={inserted}, total_source={len(src_data)}")
    out = io.BytesIO()
    wb_tgt.save(out)
    return out.getvalue()


def main() -> None:
    print(f"Download SOURCE: {DISK_SOURCE_PATH}")
    src = disk_download(DISK_SOURCE_PATH)
    print(f"downloaded SOURCE: {len(src)} bytes")

    print(f"Download TARGET: {DISK_TARGET_PATH}")
    tgt = disk_download(DISK_TARGET_PATH)
    print(f"downloaded TARGET: {len(tgt)} bytes")

    print("Run sync SVOD bools -> TARGET...")
    out_tgt = sync(src, tgt)

    print(f"Upload TARGET back: {DISK_TARGET_PATH}")
    disk_upload(DISK_TARGET_PATH, out_tgt)

    print("✅ Done")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        sys.exit(1)
