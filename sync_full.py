"""
Скрипт 1: ПОЛНАЯ СИНХРОНИЗАЦИЯ

Последовательность:
1. БД → СВОДНАЯ (внутри ВНУТРЯНКИ)
2. СВОДНАЯ → Лист1 (ВНУТРЯНКА → ВНЕШНЯЯ): ЮЛ, Terminal ID, Добавлен сертификат, Билеты продаются
3. Лист1 → СВОДНАЯ (ВНЕШНЯЯ → ВНУТРЯНКА): только «Добавлен сертификат (МТС)»
4. БД → терминалы (ВНУТРЯНКА → ВНЕШНЯЯ): каждая строка БД отдельно
"""

import io
import os
import sys
import time
from typing import Dict, List, Optional, Set

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from sync_utils import (
    disk_download, disk_upload,
    header_index_map, last_header_col, get_cell_str, is_empty_cell, get_last_data_row,
    col_to_letter, copy_row_style, copy_cell_style, ensure_columns_at_end,
    normalize_bool_to_01, apply_bool_cf,
    parse_terminal_id, compress_ranges, format_ranges,
    cert_value_from_bd_comment, normalize_mts_id, ru_to_translit,
)


# =======================
# ENV
# =======================
YANDEX_OAUTH_TOKEN = os.getenv("YANDEX_OAUTH_TOKEN", "").strip()
DISK_SOURCE_PATH = os.getenv("DISK_SOURCE_PATH", "").strip()
DISK_TARGET_PATH = os.getenv("DISK_TARGET_PATH", "").strip()

if not YANDEX_OAUTH_TOKEN:
    raise RuntimeError("ERROR: YANDEX_OAUTH_TOKEN is empty (set it in GitHub Secrets)")
if not DISK_SOURCE_PATH:
    raise RuntimeError("ERROR: DISK_SOURCE_PATH is empty (set it in GitHub Secrets)")
if not DISK_TARGET_PATH:
    raise RuntimeError("ERROR: DISK_TARGET_PATH is empty (set it in GitHub Secrets)")


# =======================
# CONFIG
# =======================
SHEET_BD = "БД"
SHEET_SVOD = "СВОДНАЯ"
SHEET_TARGET_LIST = "Лист1"
SHEET_TARGET_TERMINALS = "терминалы"

BD_CERT_COMMENT_COL = "Комментарии"

KEY_COL = "ЮЛ"
MTS_CERT_COL = "Добавлен сертификат (МТС)"
ENG_COL = "ENG"

# Колонки СВОДНОЙ
SVOD_BOOL_COLS = [
    "Добавлен сертификат",
    MTS_CERT_COL,
    "Билеты продаются",
]

SVOD_REQUIRED_BASE = [
    "ЮЛ",
    "МТС ID",
    "Terminal ID (Столото)",
    "Агент ID (Столото)",
    "GUID",
    "Ответственный ССПС",
]

BD_REQUIRED = [
    "ЮЛ",
    "МТС ID",
    "Terminal ID (Столото)",
    "Агент ID (Столото)",
    "GUID",
    "Ответственный ССПС",
]

# Колонки для синка СВОДНАЯ → Лист1 (без МТС ID!)
COLS_SYNC_TO_TARGET = [
    "ЮЛ",
    "Terminal ID (Столото)",
    "Добавлен сертификат",
    "Билеты продаются",
]

# Колонки для CF в Лист1
TARGET_COLS_WITH_CF = ["Добавлен сертификат", MTS_CERT_COL, "Билеты продаются"]

# Колонки листа «терминалы»
COL_MTS = "МТС ID"
COL_TERMINAL = "Terminal ID (Столото)"
COL_REGION = "Регион"
COL_CITY = "Город"
COL_STREET = "Улица"
COL_HOUSE = "Дом"
COL_AGENT = "Агент ID (Столото)"
COL_CERT = "Добавлен сертификат"
COL_CERT_MTS = MTS_CERT_COL
COL_COMMENTS_MTS = "Комментарии (МТС)"
COL_COMMENTS_STOLOTO = "Комментарии (Столото)"

TERMINALS_BASE_COLS = [
    KEY_COL, COL_MTS, COL_TERMINAL,
    COL_REGION, COL_CITY, COL_STREET, COL_HOUSE,
    COL_AGENT, COL_CERT,
    COL_CERT_MTS, COL_COMMENTS_MTS, COL_COMMENTS_STOLOTO,
]


# ===============================================================
# STEP 1: БД → СВОДНАЯ (внутри ВНУТРЯНКИ)
# ===============================================================
def delete_missing_uls(ws_svod: Worksheet, sv_map: Dict[str, int], uls_in_bd: Set[str]) -> int:
    ul_col = sv_map["ЮЛ"]
    last_data = get_last_data_row(ws_svod, ul_col, start_row=2)
    if last_data < 2:
        return 0
    to_delete: List[int] = []
    for r in range(2, last_data + 1):
        ul = get_cell_str(ws_svod, r, ul_col)
        if ul and ul not in uls_in_bd:
            to_delete.append(r)
    for r in reversed(to_delete):
        ws_svod.delete_rows(r, 1)
    return len(to_delete)


def step1_bd_to_svod(wb_src):
    """БД → СВОДНАЯ внутри одного workbook. Возвращает изменённый wb_src."""
    if SHEET_BD not in wb_src.sheetnames:
        raise RuntimeError(f'Source: sheet "{SHEET_BD}" not found')
    if SHEET_SVOD not in wb_src.sheetnames:
        raise RuntimeError(f'Source: sheet "{SHEET_SVOD}" not found')

    ws_bd = wb_src[SHEET_BD]
    ws_svod = wb_src[SHEET_SVOD]

    ensure_columns_at_end(ws_svod, SVOD_BOOL_COLS)

    bd_map = header_index_map(ws_bd)
    sv_map = header_index_map(ws_svod)

    missing_bd = [c for c in BD_REQUIRED if c not in bd_map]
    if missing_bd:
        raise RuntimeError(f'Missing columns in "{SHEET_BD}": {missing_bd}')
    missing_svod = [c for c in SVOD_REQUIRED_BASE if c not in sv_map]
    if missing_svod:
        raise RuntimeError(f'Missing columns in "{SHEET_SVOD}": {missing_svod}')

    ul_col_bd = bd_map["ЮЛ"]
    terminal_col_bd = bd_map["Terminal ID (Столото)"]

    bd_by_ul: Dict[str, Dict[str, str]] = {}
    terminals_by_ul: Dict[str, List[int]] = {}
    uls_in_bd: Set[str] = set()

    for r in range(2, ws_bd.max_row + 1):
        ul = get_cell_str(ws_bd, r, ul_col_bd)
        if not ul:
            continue
        uls_in_bd.add(ul)

        term_raw = ws_bd.cell(row=r, column=terminal_col_bd).value
        term_num = parse_terminal_id(term_raw) if term_raw is not None else None
        if term_num is not None:
            terminals_by_ul.setdefault(ul, []).append(term_num)

        payload = bd_by_ul.setdefault(ul, {k: "" for k in BD_REQUIRED})
        for col_name in BD_REQUIRED:
            val = get_cell_str(ws_bd, r, bd_map[col_name])
            if payload[col_name] == "" and val != "":
                payload[col_name] = val

    for ul, nums in terminals_by_ul.items():
        bd_by_ul[ul]["Terminal ID (Столото)"] = format_ranges(compress_ranges(nums))

    deleted = delete_missing_uls(ws_svod, sv_map, uls_in_bd)
    if deleted:
        print(f"  Deleted from SVOD (not in BD): {deleted}")

    sv_map = header_index_map(ws_svod)
    ul_col_sv = sv_map["ЮЛ"]
    last_data_row = get_last_data_row(ws_svod, ul_col_sv, start_row=2)
    template_row = 2 if ws_svod.max_row >= 2 else (last_data_row if last_data_row >= 2 else 2)
    max_col = ws_svod.max_column

    existing_row_by_ul: Dict[str, int] = {}
    if last_data_row >= 2:
        for r in range(2, last_data_row + 1):
            ul = get_cell_str(ws_svod, r, ul_col_sv)
            if ul:
                existing_row_by_ul[ul] = r

    inserted = 0
    updated = 0
    append_row = last_data_row + 1 if last_data_row >= 2 else 2

    for ul, payload in bd_by_ul.items():
        if ul in existing_row_by_ul:
            rr = existing_row_by_ul[ul]
            for col_name in SVOD_REQUIRED_BASE:
                ws_svod.cell(row=rr, column=sv_map[col_name]).value = payload.get(col_name, "")
            # НЕ ТРОГАЕМ: "Добавлен сертификат", "Добавлен сертификат (МТС)", "Билеты продаются"
            updated += 1
        else:
            rr = append_row
            append_row += 1
            if 2 <= template_row <= ws_svod.max_row:
                copy_row_style(ws_svod, template_row, rr, max_col)
            for col_name in SVOD_REQUIRED_BASE:
                ws_svod.cell(row=rr, column=sv_map[col_name]).value = payload.get(col_name, "")
            # все bool-колонки для новых строк: 0 по умолчанию
            ws_svod.cell(row=rr, column=sv_map["Добавлен сертификат"]).value = 0
            ws_svod.cell(row=rr, column=sv_map[MTS_CERT_COL]).value = 0
            ws_svod.cell(row=rr, column=sv_map["Билеты продаются"]).value = 0
            inserted += 1

    # нормализация 0/1
    last_data_row = get_last_data_row(ws_svod, ul_col_sv, start_row=2)
    for col_name in SVOD_BOOL_COLS:
        c = sv_map[col_name]
        for r in range(2, last_data_row + 1):
            v = ws_svod.cell(row=r, column=c).value
            if is_empty_cell(v):
                continue
            norm = normalize_bool_to_01(v)
            if norm is not None:
                ws_svod.cell(row=r, column=c).value = norm

    # CF
    for col_name in SVOD_BOOL_COLS:
        c = sv_map[col_name]
        letter = col_to_letter(c)
        apply_bool_cf(ws_svod, letter, start_row=2, end_row=max(last_data_row, 2))

    print(f"  Step 1 done: inserted={inserted}, updated={updated}, deleted={deleted}, uls={len(bd_by_ul)}")


# ===============================================================
# STEP 2: СВОДНАЯ → Лист1 (ВНУТРЯНКА → ВНЕШНЯЯ)
# ===============================================================
def step2_svod_to_target_list(wb_src, wb_tgt):
    """СВОДНАЯ → Лист1: ЮЛ, Terminal ID, Добавлен сертификат, Билеты продаются. НЕ МТС ID. НЕ ТРОГАТЬ МТС cert."""
    ws_src = wb_src[SHEET_SVOD]

    ws_tgt = (
        wb_tgt[SHEET_TARGET_LIST]
        if SHEET_TARGET_LIST in wb_tgt.sheetnames
        else wb_tgt.create_sheet(SHEET_TARGET_LIST)
    )

    src_map = header_index_map(ws_src)
    tgt_map = header_index_map(ws_tgt)

    if KEY_COL not in src_map:
        raise RuntimeError(f'Source sheet "{SHEET_SVOD}": key column "{KEY_COL}" not found')

    # Ensure headers in TARGET
    h_last = last_header_col(ws_tgt)

    def ensure_header(name: str):
        nonlocal h_last, tgt_map
        if name in tgt_map:
            return
        h_last += 1
        ws_tgt.cell(row=1, column=h_last).value = name
        tgt_map[name] = h_last

    for name in COLS_SYNC_TO_TARGET:
        ensure_header(name)
    ensure_header(ENG_COL)
    ensure_header(MTS_CERT_COL)
    tgt_map = header_index_map(ws_tgt)

    src_last = get_last_data_row(ws_src, src_map[KEY_COL], start_row=2)
    tgt_last = get_last_data_row(ws_tgt, tgt_map[KEY_COL], start_row=2)

    # Read SOURCE
    src_data: Dict[str, Dict[str, str]] = {}
    for r in range(2, src_last + 1):
        key = get_cell_str(ws_src, r, src_map[KEY_COL])
        if not key:
            continue
        row_payload: Dict[str, str] = {}
        for col in COLS_SYNC_TO_TARGET:
            row_payload[col] = get_cell_str(ws_src, r, src_map[col]) if col in src_map else ""
        src_data[key] = row_payload

    # Existing TARGET rows
    tgt_row_by_key: Dict[str, int] = {}
    if tgt_last >= 2:
        for r in range(2, tgt_last + 1):
            key = get_cell_str(ws_tgt, r, tgt_map[KEY_COL])
            if key:
                tgt_row_by_key[key] = r

    template_row = 2 if ws_tgt.max_row >= 2 else 2
    max_col = last_header_col(ws_tgt)

    updated = 0
    inserted = 0
    append_row = tgt_last + 1 if tgt_last >= 2 else 2

    for key, payload in src_data.items():
        if key in tgt_row_by_key:
            rr = tgt_row_by_key[key]
            for col in COLS_SYNC_TO_TARGET:
                ws_tgt.cell(row=rr, column=tgt_map[col]).value = payload.get(col, "")
            # MTS_CERT_COL — НЕ ТРОГАЕМ
            updated += 1
        else:
            rr = append_row
            append_row += 1
            if 2 <= template_row <= ws_tgt.max_row:
                copy_row_style(ws_tgt, template_row, rr, max_col)
            for col in COLS_SYNC_TO_TARGET:
                ws_tgt.cell(row=rr, column=tgt_map[col]).value = payload.get(col, "")
            # MTS_CERT_COL для новых — оставляем пустым (поле франчайзи)
            ws_tgt.cell(row=rr, column=tgt_map[ENG_COL]).value = None
            inserted += 1

    # Нормализация bool в TARGET
    tgt_last = max(get_last_data_row(ws_tgt, tgt_map[KEY_COL], start_row=2), 2)

    for b in ("Добавлен сертификат", "Билеты продаются"):
        if b not in tgt_map:
            continue
        c = tgt_map[b]
        for r in range(2, tgt_last + 1):
            v = ws_tgt.cell(row=r, column=c).value
            if is_empty_cell(v):
                ws_tgt.cell(row=r, column=c).value = 0
                continue
            norm = normalize_bool_to_01(v)
            if norm is not None:
                ws_tgt.cell(row=r, column=c).value = norm

    # MTS cert нормализация (не пустые)
    if MTS_CERT_COL in tgt_map:
        c = tgt_map[MTS_CERT_COL]
        for r in range(2, tgt_last + 1):
            v = ws_tgt.cell(row=r, column=c).value
            if is_empty_cell(v):
                continue
            norm = normalize_bool_to_01(v)
            if norm is not None:
                ws_tgt.cell(row=r, column=c).value = norm

    # CF
    for b in TARGET_COLS_WITH_CF:
        if b not in tgt_map:
            continue
        letter = col_to_letter(tgt_map[b])
        apply_bool_cf(ws_tgt, letter, start_row=2, end_row=tgt_last)

    # ENG autotranslit
    if KEY_COL in tgt_map and ENG_COL in tgt_map:
        ul_c = tgt_map[KEY_COL]
        eng_c = tgt_map[ENG_COL]
        last_ul = get_last_data_row(ws_tgt, ul_c, start_row=2)
        filled = 0
        for r in range(2, last_ul + 1):
            ul_val = ws_tgt.cell(row=r, column=ul_c).value
            eng_val = ws_tgt.cell(row=r, column=eng_c).value
            if is_empty_cell(ul_val):
                continue
            if not is_empty_cell(eng_val):
                continue
            ws_tgt.cell(row=r, column=eng_c).value = ru_to_translit(str(ul_val))
            filled += 1
        if filled:
            print(f"  ENG filled: {filled}")

    print(f"  Step 2 done: updated={updated}, inserted={inserted}, total={len(src_data)}")


# ===============================================================
# STEP 3: Лист1 → СВОДНАЯ (только «Добавлен сертификат (МТС)»)
# ===============================================================
def step3_target_to_svod_mts_cert(wb_src, wb_tgt):
    """Лист1 → СВОДНАЯ: переносим только «Добавлен сертификат (МТС)»."""
    ws_src = wb_src[SHEET_SVOD]
    ws_tgt = wb_tgt[SHEET_TARGET_LIST]

    src_map = header_index_map(ws_src)
    tgt_map = header_index_map(ws_tgt)

    if KEY_COL not in src_map:
        raise RuntimeError(f'SOURCE СВОДНАЯ: key column "{KEY_COL}" not found')
    if KEY_COL not in tgt_map:
        raise RuntimeError(f'TARGET Лист1: key column "{KEY_COL}" not found')

    src_key_c = src_map[KEY_COL]
    tgt_key_c = tgt_map[KEY_COL]

    # Ensure col in SOURCE
    if MTS_CERT_COL not in src_map:
        ensure_columns_at_end(ws_src, [MTS_CERT_COL])
        src_map = header_index_map(ws_src)
    src_col_c = src_map[MTS_CERT_COL]

    if MTS_CERT_COL not in tgt_map:
        print("  Step 3 skip: MTS cert column not found in TARGET")
        return

    tgt_col_c = tgt_map[MTS_CERT_COL]

    # Build dict from TARGET
    tgt_last = get_last_data_row(ws_tgt, tgt_key_c, start_row=2)
    data: Dict[str, int] = {}
    for r in range(2, tgt_last + 1):
        key = ws_tgt.cell(row=r, column=tgt_key_c).value
        if is_empty_cell(key):
            continue
        v = ws_tgt.cell(row=r, column=tgt_col_c).value
        norm = normalize_bool_to_01(v)
        if norm is None:
            continue
        data[str(key).strip()] = norm

    # Apply to SOURCE
    src_last = get_last_data_row(ws_src, src_key_c, start_row=2)
    updated = 0
    for r in range(2, src_last + 1):
        key = ws_src.cell(row=r, column=src_key_c).value
        if is_empty_cell(key):
            continue
        k = str(key).strip()
        if k not in data:
            continue
        ws_src.cell(row=r, column=src_col_c).value = data[k]
        updated += 1

    # CF
    letter = col_to_letter(src_col_c)
    apply_bool_cf(ws_src, letter, start_row=2, end_row=max(src_last, 2))

    print(f"  Step 3 done: updated={updated}, keys_with_value={len(data)}")


# ===============================================================
# STEP 4: БД → терминалы (ВНУТРЯНКА → ВНЕШНЯЯ)
# ===============================================================
def pick_bd_mts_col(bd_map: Dict[str, int]) -> Optional[str]:
    if COL_MTS in bd_map:
        return COL_MTS
    if "МТСID" in bd_map:
        return "МТСID"
    return None


def step4_bd_to_terminals(wb_src, wb_tgt):
    """БД → терминалы: каждая строка БД без группировки."""
    ws_bd = wb_src[SHEET_BD]

    ws_tgt = (
        wb_tgt[SHEET_TARGET_TERMINALS]
        if SHEET_TARGET_TERMINALS in wb_tgt.sheetnames
        else wb_tgt.create_sheet(SHEET_TARGET_TERMINALS)
    )

    bd_map = header_index_map(ws_bd)

    if COL_AGENT not in bd_map:
        raise RuntimeError(f'BD missing required column: "{COL_AGENT}"')
    if COL_TERMINAL not in bd_map:
        raise RuntimeError(f'BD missing required column: "{COL_TERMINAL}"')

    bd_mts_name = pick_bd_mts_col(bd_map)
    bd_has_comments = BD_CERT_COMMENT_COL in bd_map

    ensure_columns_at_end(ws_tgt, TERMINALS_BASE_COLS)
    tgt_map = header_index_map(ws_tgt)

    bd_last = get_last_data_row(ws_bd, bd_map[COL_AGENT], start_row=2)
    tgt_last = get_last_data_row(ws_tgt, tgt_map.get(COL_AGENT, 1), start_row=2) if COL_AGENT in tgt_map else 1

    row_by_agent: Dict[str, int] = {}
    if COL_AGENT in tgt_map and tgt_last >= 2:
        for r in range(2, tgt_last + 1):
            a = get_cell_str(ws_tgt, r, tgt_map[COL_AGENT])
            if a:
                row_by_agent[a] = r

    template_row = 2 if ws_tgt.max_row >= 2 else (tgt_last if tgt_last >= 2 else 2)
    max_col = last_header_col(ws_tgt)

    updated = 0
    inserted = 0

    SYNC_COLS = [KEY_COL, COL_MTS, COL_TERMINAL, COL_REGION, COL_CITY, COL_STREET, COL_HOUSE, COL_AGENT, COL_CERT]

    for r in range(2, bd_last + 1):
        agent = get_cell_str(ws_bd, r, bd_map[COL_AGENT])
        if not agent:
            continue

        bd_comment_val = ws_bd.cell(row=r, column=bd_map[BD_CERT_COMMENT_COL]).value if bd_has_comments else None
        cert_val = cert_value_from_bd_comment(bd_comment_val)

        def bd_val(col_name: str) -> str:
            if col_name == COL_MTS:
                if bd_mts_name and bd_mts_name in bd_map:
                    return normalize_mts_id(ws_bd.cell(row=r, column=bd_map[bd_mts_name]).value)
                return ""
            if col_name in bd_map:
                return get_cell_str(ws_bd, r, bd_map[col_name])
            return ""

        payload: Dict[str, object] = {
            KEY_COL: bd_val(KEY_COL),
            COL_MTS: bd_val(COL_MTS),
            COL_TERMINAL: bd_val(COL_TERMINAL),
            COL_REGION: bd_val(COL_REGION),
            COL_CITY: bd_val(COL_CITY),
            COL_STREET: bd_val(COL_STREET),
            COL_HOUSE: bd_val(COL_HOUSE),
            COL_AGENT: agent,
            COL_CERT: cert_val,
        }

        if agent in row_by_agent:
            rr = row_by_agent[agent]
            for col in SYNC_COLS:
                if col in tgt_map:
                    ws_tgt.cell(row=rr, column=tgt_map[col]).value = payload.get(col, "")
            # НЕ ТРОГАЕМ: COL_CERT_MTS, COL_COMMENTS_MTS, COL_COMMENTS_STOLOTO
            updated += 1
        else:
            rr = max(tgt_last, 1) + 1
            tgt_last = rr
            row_by_agent[agent] = rr

            if 2 <= template_row <= ws_tgt.max_row:
                copy_row_style(ws_tgt, template_row, rr, max_col)

            for col in SYNC_COLS:
                if col in tgt_map:
                    ws_tgt.cell(row=rr, column=tgt_map[col]).value = payload.get(col, "")

            if COL_CERT_MTS in tgt_map:
                ws_tgt.cell(row=rr, column=tgt_map[COL_CERT_MTS]).value = 0
            if COL_COMMENTS_MTS in tgt_map:
                ws_tgt.cell(row=rr, column=tgt_map[COL_COMMENTS_MTS]).value = ""
            if COL_COMMENTS_STOLOTO in tgt_map:
                ws_tgt.cell(row=rr, column=tgt_map[COL_COMMENTS_STOLOTO]).value = ""
            inserted += 1

    # CF
    end_row = max(get_last_data_row(ws_tgt, tgt_map[COL_AGENT], start_row=2) if COL_AGENT in tgt_map else 2, 2)
    for col_name in [COL_CERT, COL_CERT_MTS]:
        if col_name in tgt_map:
            letter = col_to_letter(tgt_map[col_name])
            apply_bool_cf(ws_tgt, letter, start_row=2, end_row=end_row)

    print(f"  Step 4 done: updated={updated}, inserted={inserted}, bd_rows={max(bd_last - 1, 0)}")


# ===============================================================
# HELPERS: workbook ↔ bytes
# ===============================================================
def wb_to_bytes(wb) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ===============================================================
# ENTRYPOINT
# ===============================================================
def main() -> None:
    print(f"Download SOURCE: {DISK_SOURCE_PATH}")
    src_bytes = disk_download(DISK_SOURCE_PATH, YANDEX_OAUTH_TOKEN)
    print(f"  SOURCE: {len(src_bytes)} bytes")

    print(f"Download TARGET: {DISK_TARGET_PATH}")
    tgt_bytes = disk_download(DISK_TARGET_PATH, YANDEX_OAUTH_TOKEN)
    print(f"  TARGET: {len(tgt_bytes)} bytes")

    wb_src = load_workbook(io.BytesIO(src_bytes))
    wb_tgt = load_workbook(io.BytesIO(tgt_bytes))

    # Step 1: БД → СВОДНАЯ
    print("Step 1: БД → СВОДНАЯ...")
    step1_bd_to_svod(wb_src)

    # Step 2: СВОДНАЯ → Лист1
    print("Step 2: СВОДНАЯ → Лист1...")
    step2_svod_to_target_list(wb_src, wb_tgt)

    # Step 3: Лист1 → СВОДНАЯ (МТС cert)
    print("Step 3: Лист1 → СВОДНАЯ (МТС cert)...")
    step3_target_to_svod_mts_cert(wb_src, wb_tgt)

    # Step 4: БД → терминалы
    print("Step 4: БД → терминалы...")
    step4_bd_to_terminals(wb_src, wb_tgt)

    # Upload обоих файлов
    print(f"Upload SOURCE: {DISK_SOURCE_PATH}")
    disk_upload(DISK_SOURCE_PATH, wb_to_bytes(wb_src), YANDEX_OAUTH_TOKEN)

    print(f"Upload TARGET: {DISK_TARGET_PATH}")
    disk_upload(DISK_TARGET_PATH, wb_to_bytes(wb_tgt), YANDEX_OAUTH_TOKEN)

    print("Done: full sync")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
