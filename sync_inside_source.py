"""
Скрипт 2: ВНУТРЯНКА.БД → ВНУТРЯНКА.СВОДНАЯ

- Группирует строки БД по ЮЛ (один ЮЛ = одна строка СВОДНОЙ)
- Terminal ID → диапазоны: (31954552) или (31954552–31954557) или (31954662–31954680) (31954817)
- Добавлен сертификат → вычисляется по комментарию из БД
- НЕ ТРОГАЕТ: «Добавлен сертификат (МТС)», «Билеты продаются» у существующих строк
"""

import io
import os
import sys
from typing import Dict, List, Set

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.worksheet.datavalidation import DataValidation

from sync_utils import (
    disk_download, disk_upload,
    header_index_map, get_cell_str, is_empty_cell, get_last_data_row,
    col_to_letter, copy_row_style, ensure_columns_at_end,
    normalize_bool_to_01, apply_bool_cf,
    parse_terminal_id, compress_ranges, format_ranges,
    COL_STATUS, STATUS_VALUES, STATUS_DEFAULT,
)


# =======================
# ENV
# =======================
YANDEX_OAUTH_TOKEN = os.getenv("YANDEX_OAUTH_TOKEN", "").strip()
DISK_SOURCE_PATH = os.getenv("DISK_SOURCE_PATH", "").strip()

if not YANDEX_OAUTH_TOKEN:
    raise RuntimeError("ERROR: YANDEX_OAUTH_TOKEN is empty (set it in GitHub Secrets)")
if not DISK_SOURCE_PATH:
    raise RuntimeError("ERROR: DISK_SOURCE_PATH is empty (set it in GitHub Secrets)")


# =======================
# CONFIG
# =======================
SHEET_BD = "БД"
SHEET_SVOD = "СВОДНАЯ"

SVOD_BOOL_COLS = [
    "Добавлен сертификат",
    "Добавлен сертификат (МТС)",
    "Билеты продаются",
]

SVOD_REQUIRED_BASE = [
    "ЮЛ",
    "МТС ID",
    "Terminal ID (Столото)",
    "Агент ID (Столото)",
    "GUID",
    "Ответственный ССПС",
]

BD_REQUIRED = [
    "ЮЛ",
    "МТС ID",
    "Terminal ID (Столото)",
    "Агент ID (Столото)",
    "GUID",
    "Ответственный ССПС",
]


# =======================
# DELETE ЮЛ removed from BD
# =======================
def delete_missing_uls(ws_svod: Worksheet, sv_map: Dict[str, int], uls_in_bd: Set[str]) -> int:
    ul_col = sv_map["ЮЛ"]
    last_data = get_last_data_row(ws_svod, ul_col, start_row=2)
    if last_data < 2:
        return 0

    to_delete: List[int] = []
    for r in range(2, last_data + 1):
        ul = get_cell_str(ws_svod, r, ul_col)
        if ul and ul not in uls_in_bd:
            to_delete.append(r)

    deleted = 0
    for r in reversed(to_delete):
        ws_svod.delete_rows(r, 1)
        deleted += 1
    return deleted


# =======================
# MAIN SYNC LOGIC
# =======================
def sync_inside_workbook(src_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(src_bytes))

    if SHEET_BD not in wb.sheetnames:
        raise RuntimeError(f'Source: sheet "{SHEET_BD}" not found')
    if SHEET_SVOD not in wb.sheetnames:
        raise RuntimeError(f'Target: sheet "{SHEET_SVOD}" not found')

    ws_bd = wb[SHEET_BD]
    ws_svod = wb[SHEET_SVOD]

    # Добавляем колонку "Статус" в БД, если её нет, + data validation
    ensure_columns_at_end(ws_bd, [COL_STATUS])
    _bd_map_tmp = header_index_map(ws_bd)
    _status_letter = col_to_letter(_bd_map_tmp[COL_STATUS])
    _bd_last_tmp = get_last_data_row(ws_bd, _bd_map_tmp.get("ЮЛ", 1), start_row=2)
    _dv = DataValidation(
        type="list",
        formula1='"' + ",".join(STATUS_VALUES) + '"',
        allow_blank=True,
    )
    _dv.error = "Выберите значение из списка"
    _dv.errorTitle = "Ошибка"
    _dv.prompt = "Выберите статус"
    _dv.promptTitle = "Статус"
    ws_bd.add_data_validation(_dv)
    _dv.add(f"{_status_letter}2:{_status_letter}{max(_bd_last_tmp, 2) + 500}")

    # Заполняем пустые ячейки Статус значением по умолчанию
    _status_col = _bd_map_tmp[COL_STATUS]
    for r in range(2, _bd_last_tmp + 1):
        v = ws_bd.cell(row=r, column=_status_col).value
        if v is None or str(v).strip() == "":
            ws_bd.cell(row=r, column=_status_col).value = STATUS_DEFAULT

    print(f'Ensure columns in "{SHEET_SVOD}"...')
    ensure_columns_at_end(ws_svod, SVOD_BOOL_COLS)

    bd_map = header_index_map(ws_bd)
    sv_map = header_index_map(ws_svod)

    missing_bd = [c for c in BD_REQUIRED if c not in bd_map]
    if missing_bd:
        raise RuntimeError(f'Missing columns in "{SHEET_BD}": {missing_bd}')

    missing_svod = [c for c in SVOD_REQUIRED_BASE if c not in sv_map]
    if missing_svod:
        raise RuntimeError(f'Missing columns in "{SHEET_SVOD}": {missing_svod}')

    # Ключ группировки — ЮЛ (один ЮЛ = одна строка СВОДНОЙ)
    ul_col_bd = bd_map["ЮЛ"]
    terminal_col_bd = bd_map["Terminal ID (Столото)"]

    bd_by_ul: Dict[str, Dict[str, str]] = {}
    terminals_by_ul: Dict[str, List[int]] = {}
    uls_in_bd: Set[str] = set()

    for r in range(2, ws_bd.max_row + 1):
        ul = get_cell_str(ws_bd, r, ul_col_bd)
        if not ul:
            continue

        uls_in_bd.add(ul)

        term_raw = ws_bd.cell(row=r, column=terminal_col_bd).value
        term_num = parse_terminal_id(term_raw) if term_raw is not None else None
        if term_num is not None:
            terminals_by_ul.setdefault(ul, []).append(term_num)

        payload = bd_by_ul.setdefault(ul, {k: "" for k in BD_REQUIRED})
        for col_name in BD_REQUIRED:
            val = get_cell_str(ws_bd, r, bd_map[col_name])
            if payload[col_name] == "" and val != "":
                payload[col_name] = val

    for ul, nums in terminals_by_ul.items():
        rngs = compress_ranges(nums)
        bd_by_ul[ul]["Terminal ID (Столото)"] = format_ranges(rngs)

    deleted = delete_missing_uls(ws_svod, sv_map, uls_in_bd)
    if deleted:
        print(f"Deleted from SVOD (not in BD): {deleted}")

    sv_map = header_index_map(ws_svod)
    ul_col_sv = sv_map["ЮЛ"]

    last_data_row = get_last_data_row(ws_svod, ul_col_sv, start_row=2)

    template_row = 2 if ws_svod.max_row >= 2 else (last_data_row if last_data_row >= 2 else 2)
    max_col = ws_svod.max_column

    existing_row_by_ul: Dict[str, int] = {}
    if last_data_row >= 2:
        for r in range(2, last_data_row + 1):
            ul = get_cell_str(ws_svod, r, ul_col_sv)
            if ul:
                existing_row_by_ul[ul] = r

    inserted = 0
    updated = 0
    append_row = last_data_row + 1 if last_data_row >= 2 else 2

    for ul, payload in bd_by_ul.items():
        if ul in existing_row_by_ul:
            rr = existing_row_by_ul[ul]
            for col_name in SVOD_REQUIRED_BASE:
                ws_svod.cell(row=rr, column=sv_map[col_name]).value = payload.get(col_name, "")
            # НЕ ТРОГАЕМ: "Добавлен сертификат", "Добавлен сертификат (МТС)", "Билеты продаются"
            updated += 1
        else:
            rr = append_row
            append_row += 1

            if 2 <= template_row <= ws_svod.max_row:
                copy_row_style(ws_svod, template_row, rr, max_col)

            for col_name in SVOD_REQUIRED_BASE:
                ws_svod.cell(row=rr, column=sv_map[col_name]).value = payload.get(col_name, "")
            # все bool-колонки для новых строк: 0 по умолчанию
            ws_svod.cell(row=rr, column=sv_map["Добавлен сертификат"]).value = 0
            ws_svod.cell(row=rr, column=sv_map["Добавлен сертификат (МТС)"]).value = 0
            ws_svod.cell(row=rr, column=sv_map["Билеты продаются"]).value = 0
            inserted += 1

    # нормализация 0/1
    last_data_row = get_last_data_row(ws_svod, ul_col_sv, start_row=2)
    for col_name in SVOD_BOOL_COLS:
        c = sv_map[col_name]
        for r in range(2, last_data_row + 1):
            v = ws_svod.cell(row=r, column=c).value
            if is_empty_cell(v):
                continue
            norm = normalize_bool_to_01(v)
            if norm is not None:
                ws_svod.cell(row=r, column=c).value = norm

    # CF на реальные строки данных
    for col_name in SVOD_BOOL_COLS:
        c = sv_map[col_name]
        letter = col_to_letter(c)
        apply_bool_cf(ws_svod, letter, start_row=2, end_row=max(last_data_row, 2))

    print(
        f"Inside sync done: inserted={inserted}, updated={updated}, deleted={deleted}, "
        f"total_source_uls={len(bd_by_ul)}"
    )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =======================
# ENTRYPOINT
# =======================
def main() -> None:
    print(f"Download SOURCE: {DISK_SOURCE_PATH}")
    src = disk_download(DISK_SOURCE_PATH, YANDEX_OAUTH_TOKEN)
    print(f"downloaded SOURCE: {len(src)} bytes")

    print("Running inside SOURCE sync (БД → СВОДНАЯ)...")
    out = sync_inside_workbook(src)

    print(f"Upload back to same path (SOURCE): {DISK_SOURCE_PATH}")
    disk_upload(DISK_SOURCE_PATH, out, YANDEX_OAUTH_TOKEN)
    print("Done: inside SOURCE")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
