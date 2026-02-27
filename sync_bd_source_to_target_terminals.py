import io
import os
import sys
import time
from typing import Dict, List, Optional, Set, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from copy import copy


# =======================
# ENV (НЕ ПЕРЕИМЕНОВЫВАТЬ)
# =======================
YANDEX_OAUTH_TOKEN = os.getenv("YANDEX_OAUTH_TOKEN", "").strip()
DISK_SOURCE_PATH = os.getenv("DISK_SOURCE_PATH", "").strip()
DISK_TARGET_PATH = os.getenv("DISK_TARGET_PATH", "").strip()  # НЕ УДАЛЯТЬ

if not YANDEX_OAUTH_TOKEN:
    raise RuntimeError("ERROR: YANDEX_OAUTH_TOKEN is empty (set it in GitHub Secrets)")
if not DISK_SOURCE_PATH:
    raise RuntimeError("ERROR: DISK_SOURCE_PATH is empty (set it in GitHub Secrets)")
if not DISK_TARGET_PATH:
    raise RuntimeError("ERROR: DISK_TARGET_PATH is empty (set it in GitHub Secrets)")

YANDEX_API = "https://cloud-api.yandex.net/v1/disk"
HEADERS = {"Authorization": f"OAuth {YANDEX_OAUTH_TOKEN}"}


# =======================
# CONFIG
# =======================
SRC_BD_SHEET = "БД"
TGT_SHEET = "терминалы"

COL_UL = "ЮЛ"
COL_MTS = "МТС ID"  # в БД может быть "МТСID" — обработаем ниже
COL_TERMINAL = "Terminal ID (Столото)"
COL_REGION = "Регион"
COL_CITY = "Город"
COL_STREET = "Улица"
COL_HOUSE = "Дом"
COL_AGENT = "Агент ID (Столото)"

# В БД комментарии обычно в колонке "Комментарии"
COL_BD_COMMENTS = "Комментарии"

# В TARGET сохраняем существующее и НЕ ЗАТИРАЕМ:
COL_CERT_MTS = "Добавлен сертификат (МТС)"
COL_COMMENTS_MTS = "Комментарии (МТС)"
COL_COMMENTS_STOLOTO = "Комментарии (Столото)"

# Меняем ТОЛЬКО это поле по условию:
COL_CERT = "Добавлен сертификат"

# Столбцы, которые должны быть в TARGET (если нет — создать). Заполнение из БД если колонка есть, иначе пусто.
TARGET_BASE_COLS = [
    COL_UL,
    COL_MTS,
    COL_TERMINAL,
    COL_REGION,
    COL_CITY,
    COL_STREET,
    COL_HOUSE,
    COL_AGENT,
    COL_CERT,
    COL_CERT_MTS,
    "Комментарии",         # общий комментарий (если используешь)
    COL_COMMENTS_MTS,
    COL_COMMENTS_STOLOTO,
]

# Фраза-исключение (точное совпадение после trim+lower)
CERT_OK_PHRASE = "есть все, но со стороны мтс нет сертификата"


# =======================
# YANDEX DISK API
# =======================
def disk_download(path: str) -> bytes:
    r = requests.get(
        f"{YANDEX_API}/resources/download",
        headers=HEADERS,
        params={"path": path},
        timeout=60,
    )
    if r.status_code >= 400:
        raise RuntimeError(f"DOWNLOAD ERROR: {r.status_code}\nPATH: {path}\nBODY: {r.text}")
    href = r.json()["href"]

    f = requests.get(href, timeout=180)
    if f.status_code >= 400:
        raise RuntimeError(f"DOWNLOAD(HREF) ERROR: {f.status_code}\nHREF: {href}\nBODY: {f.text}")
    return f.content


def disk_upload(path: str, content: bytes, retries: int = 8) -> None:
    r = requests.get(
        f"{YANDEX_API}/resources/upload",
        headers=HEADERS,
        params={"path": path, "overwrite": "true"},
        timeout=60,
    )
    if r.status_code >= 400:
        raise RuntimeError(f"UPLOAD(HREF) ERROR: {r.status_code}\nPATH: {path}\nBODY: {r.text}")
    href = r.json()["href"]

    for attempt in range(1, retries + 1):
        put = requests.put(href, data=content, timeout=240)
        if put.status_code < 400:
            return

        if put.status_code == 423:
            wait = min(2 ** attempt, 30)
            print(f"⚠️ Upload LOCKED (423). Retry {attempt}/{retries} in {wait}s...")
            time.sleep(wait)
            continue

        raise RuntimeError(f"UPLOAD ERROR: {put.status_code}\nPATH: {path}\nBODY: {put.text}")

    raise RuntimeError(
        "UPLOAD ERROR: file is LOCKED too long (423). "
        "Закрой файл в Яндекс Таблицах/редакторе и запусти workflow ещё раз."
    )


# =======================
# HELPERS
# =======================
def header_index_map(ws: Worksheet) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            m[name] = c
    return m


def last_header_col(ws: Worksheet) -> int:
    last = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        if str(v).strip() != "":
            last = c
    return max(last, 1)


def col_to_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def copy_cell_style(src_cell, dst_cell) -> None:
    """
    Копируем стиль через copy(), чтобы не было общих ссылок на стиль (StyleProxy).
    Иначе при сохранении может "плыть" оформление.
    """
    if not src_cell.has_style:
        return

    dst_cell._style = copy(src_cell._style)
    dst_cell.font = copy(src_cell.font)
    dst_cell.border = copy(src_cell.border)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.alignment = copy(src_cell.alignment)


def ensure_headers(ws: Worksheet, headers: List[str]) -> None:
    m = header_index_map(ws)
    col = last_header_col(ws)

    # шаблон: последняя реальная колонка заголовков
    template_col = col if col >= 1 else 1
    template_header = ws.cell(row=1, column=template_col)
    template_letter = col_to_letter(template_col)
    template_width = ws.column_dimensions[template_letter].width

    for h in headers:
        if h in m:
            continue
        col += 1

        dst_header = ws.cell(row=1, column=col)
        dst_header.value = h

        # стиль заголовка
        copy_cell_style(template_header, dst_header)

        # ширина колонки
        new_letter = col_to_letter(col)
        if template_width is not None:
            ws.column_dimensions[new_letter].width = template_width

        m[h] = col


def get_cell_str(ws: Worksheet, r: int, c: int) -> str:
    v = ws.cell(row=r, column=c).value
    return "" if v is None else str(v).strip()


def is_empty(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def get_last_data_row(ws: Worksheet, key_col: int, start_row: int = 2) -> int:
    last = 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if not is_empty(v):
            last = r
    return last


def normalize_mts_id(value) -> str:
    """
    Оставляем ведущие нули. Если в ячейке число — превратим в строку, добьём до 9 цифр.
    """
    if value is None:
        return ""
    s = str(value).strip()
    if s == "":
        return ""
    digits = "".join(ch for ch in s if ch.isdigit())
    if digits == "":
        return ""
    if len(digits) > 9:
        return digits
    return digits.zfill(9)


def cert_value_from_bd_comment(comment_value) -> int:
    """
    ТВОЁ УСЛОВИЕ (меняем только поле "Добавлен сертификат"):
    - если в БД в "Комментарии" пусто ИЛИ фраза "есть все, но со стороны мтс нет сертификата"
      => 1
    - иначе => 0
    """
    s = "" if comment_value is None else str(comment_value).strip().lower()
    if s == "":
        return 1
    if s == CERT_OK_PHRASE:
        return 1
    return 0


def copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    # высота
    try:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    except Exception:
        pass

    for c in range(1, max_col + 1):
        s = ws.cell(row=src_row, column=c)
        d = ws.cell(row=dst_row, column=c)
        copy_cell_style(s, d)


# =======================
# CF (0/1)
# =======================
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GRAY = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")


def apply_bool_cf(ws: Worksheet, col_letter: str, start_row: int, end_row: int) -> None:
    """
    Проставляет CF:
    - пусто -> серый
    - 1 -> зелёный
    - 0 -> красный
    """
    if end_row < start_row:
        end_row = start_row
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    r0 = start_row


    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'LEN(TRIM({col_letter}{r0}))=0'], fill=FILL_GRAY, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'{col_letter}{r0}=1'], fill=FILL_GREEN, stopIfTrue=False),
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(formula=[f'{col_letter}{r0}=0'], fill=FILL_RED, stopIfTrue=False),
    )


# =======================
# MAIN LOGIC
# =======================
def pick_bd_mts_col(bd_map: Dict[str, int]) -> Optional[str]:
    # иногда в БД бывает "МТСID"
    if COL_MTS in bd_map:
        return COL_MTS
    if "МТСID" in bd_map:
        return "МТСID"
    return None


def sync_bd_to_target(source_bytes: bytes, target_bytes: bytes) -> bytes:
    wb_src = load_workbook(io.BytesIO(source_bytes))
    wb_tgt = load_workbook(io.BytesIO(target_bytes))

    if SRC_BD_SHEET not in wb_src.sheetnames:
        raise RuntimeError(f'SOURCE: sheet "{SRC_BD_SHEET}" not found')
    ws_bd = wb_src[SRC_BD_SHEET]

    ws_tgt = wb_tgt[TGT_SHEET] if TGT_SHEET in wb_tgt.sheetnames else wb_tgt.create_sheet(TGT_SHEET)

    bd_map = header_index_map(ws_bd)

    # Минимум для ключа/сопоставления
    if COL_AGENT not in bd_map:
        raise RuntimeError(f'BD missing required column: "{COL_AGENT}"')
    if COL_TERMINAL not in bd_map:
        raise RuntimeError(f'BD missing required column: "{COL_TERMINAL}"')

    bd_mts_name = pick_bd_mts_col(bd_map)
    bd_has_comments = COL_BD_COMMENTS in bd_map

    # TARGET: гарантируем заголовки (создаём, но НЕ трогаем существующие данные/оформление)
    ensure_headers(ws_tgt, TARGET_BASE_COLS)
    tgt_map = header_index_map(ws_tgt)

    # Границы данных
    bd_last = get_last_data_row(ws_bd, bd_map[COL_AGENT], start_row=2)
    tgt_last = get_last_data_row(ws_tgt, tgt_map.get(COL_AGENT, 1), start_row=2) if COL_AGENT in tgt_map else 1

    # Мапа TARGET: agentId -> row
    row_by_agent: Dict[str, int] = {}
    if COL_AGENT in tgt_map and tgt_last >= 2:
        for r in range(2, tgt_last + 1):
            a = get_cell_str(ws_tgt, r, tgt_map[COL_AGENT])
            if a:
                row_by_agent[a] = r

    # Шаблон стиля для новых строк: строка 2, если есть; иначе последняя строка данных; иначе просто 2
    template_row = 2 if ws_tgt.max_row >= 2 else (tgt_last if tgt_last >= 2 else 2)
    max_col = last_header_col(ws_tgt)

    updated = 0
    inserted = 0

    # Пройдёмся по БД: 1 строка БД = 1 строка TARGET по agentId (если у агента несколько строк в БД — берём последнюю по циклу)
    for r in range(2, bd_last + 1):
        agent = get_cell_str(ws_bd, r, bd_map[COL_AGENT])
        if not agent:
            continue

        # вычисляем "Добавлен сертификат" по комментариям в БД
        bd_comment_val = ws_bd.cell(row=r, column=bd_map[COL_BD_COMMENTS]).value if bd_has_comments else None
        cert_val = cert_value_from_bd_comment(bd_comment_val)

        # Подготовим значения из БД (если колонок нет — пусто)
        def bd_val(col_name: str) -> str:
            if col_name == COL_MTS:
                if bd_mts_name and bd_mts_name in bd_map:
                    return normalize_mts_id(ws_bd.cell(row=r, column=bd_map[bd_mts_name]).value)
                return ""
            if col_name in bd_map:
                return get_cell_str(ws_bd, r, bd_map[col_name])
            return ""

        payload: Dict[str, object] = {
            COL_UL: bd_val(COL_UL),
            COL_MTS: bd_val(COL_MTS),
            COL_TERMINAL: bd_val(COL_TERMINAL),
            COL_REGION: bd_val(COL_REGION),
            COL_CITY: bd_val(COL_CITY),
            COL_STREET: bd_val(COL_STREET),
            COL_HOUSE: bd_val(COL_HOUSE),
            COL_AGENT: agent,
            COL_CERT: cert_val,
            # Остальные колонки мы НЕ берём из БД (чтобы ничего не затереть):
            # COL_CERT_MTS, COL_COMMENTS_MTS, COL_COMMENTS_STOLOTO, "Комментарии"
        }

        if agent in row_by_agent:
            rr = row_by_agent[agent]

            # Обновляем базовые поля (из БД) + "Добавлен сертификат" (по условию)
            for col in [COL_UL, COL_MTS, COL_TERMINAL, COL_REGION, COL_CITY, COL_STREET, COL_HOUSE, COL_AGENT, COL_CERT]:
                if col not in tgt_map:
                    continue
                ws_tgt.cell(row=rr, column=tgt_map[col]).value = payload.get(col, "")

            # ВАЖНО: НЕ трогаем:
            # - "Добавлен сертификат (МТС)"
            # - "Комментарии (МТС)"
            # - "Комментарии (Столото)"
            updated += 1
        else:
            # новая строка
            rr = max(tgt_last, 1) + 1
            tgt_last = rr
            row_by_agent[agent] = rr

            # стиль строки
            if template_row >= 2 and template_row <= ws_tgt.max_row:
                copy_row_style(ws_tgt, template_row, rr, max_col)

            # Заполняем базовые + cert
            for col in [COL_UL, COL_MTS, COL_TERMINAL, COL_REGION, COL_CITY, COL_STREET, COL_HOUSE, COL_AGENT, COL_CERT]:
                if col not in tgt_map:
                    continue
                ws_tgt.cell(row=rr, column=tgt_map[col]).value = payload.get(col, "")

            # Для новых строк: не синкаем столото-комментарии — оставляем пусто, и НЕ затираем (т.к. новые)
            if COL_CERT_MTS in tgt_map:
                ws_tgt.cell(row=rr, column=tgt_map[COL_CERT_MTS]).value = 0
            if COL_COMMENTS_MTS in tgt_map:
                ws_tgt.cell(row=rr, column=tgt_map[COL_COMMENTS_MTS]).value = ""
            if COL_COMMENTS_STOLOTO in tgt_map:
                ws_tgt.cell(row=rr, column=tgt_map[COL_COMMENTS_STOLOTO]).value = ""

            inserted += 1

    # Условное форматирование для двух колонок: "Добавлен сертификат" и "Добавлен сертификат (МТС)"
    end_row = max(get_last_data_row(ws_tgt, tgt_map[COL_AGENT], start_row=2) if COL_AGENT in tgt_map else 2, 2)

    for col_name in [COL_CERT, COL_CERT_MTS]:
        if col_name in tgt_map:
            letter = col_to_letter(tgt_map[col_name])
            apply_bool_cf(ws_tgt, letter, start_row=2, end_row=end_row)

    print(f"BD->TARGET done: updated={updated}, inserted={inserted}, total_bd_rows={max(bd_last-1,0)}")

    out = io.BytesIO()
    wb_tgt.save(out)
    return out.getvalue()


def main() -> None:
    print(f"Download SOURCE: {DISK_SOURCE_PATH}")
    src = disk_download(DISK_SOURCE_PATH)
    print(f"downloaded SOURCE: {len(src)} bytes")

    print(f"Download TARGET: {DISK_TARGET_PATH}")
    tgt = disk_download(DISK_TARGET_PATH)
    print(f"downloaded TARGET: {len(tgt)} bytes")

    print("Run sync BD -> TARGET/терминалы (ONLY set 'Добавлен сертификат', preserve MTS/Stoloto columns)...")
    out_tgt = sync_bd_to_target(src, tgt)

    print(f"Upload TARGET back: {DISK_TARGET_PATH}")
    disk_upload(DISK_TARGET_PATH, out_tgt)

    print("✅ Done")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"❌ ERROR: {e}")
        sys.exit(1)
