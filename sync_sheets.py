"""
Скрипт 3: СВОДНАЯ ↔ Лист1

Направления:
1. СВОДНАЯ → Лист1: только булевые (Добавлен сертификат, Билеты продаются)
2. Лист1 → СВОДНАЯ: только «Добавлен сертификат (МТС)»

НЕ синкает: ЮЛ (структуру), Terminal ID — для этого нужен sync_full.py.
НЕ ТРОГАЕТ: «Добавлен сертификат (МТС)» в Лист1 (это поле франчайзи).
"""

import io
import os
import sys
from typing import Dict, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from sync_utils import (
    disk_download, disk_upload,
    header_index_map, last_header_col, get_cell_str, is_empty_cell, get_last_data_row,
    col_to_letter, copy_row_style, ensure_columns_at_end,
    normalize_bool_to_01, apply_bool_cf,
)


# =======================
# ENV
# =======================
YANDEX_OAUTH_TOKEN = os.getenv("YANDEX_OAUTH_TOKEN", "").strip()
DISK_SOURCE_PATH = os.getenv("DISK_SOURCE_PATH", "").strip()
DISK_TARGET_PATH = os.getenv("DISK_TARGET_PATH", "").strip()

if not YANDEX_OAUTH_TOKEN:
    raise RuntimeError("ERROR: YANDEX_OAUTH_TOKEN is empty (set it in GitHub Secrets)")
if not DISK_SOURCE_PATH:
    raise RuntimeError("ERROR: DISK_SOURCE_PATH is empty (set it in GitHub Secrets)")
if not DISK_TARGET_PATH:
    raise RuntimeError("ERROR: DISK_TARGET_PATH is empty (set it in GitHub Secrets)")


# =======================
# CONFIG
# =======================
SRC_SHEET = "СВОДНАЯ"
TGT_SHEET = "Лист1"
KEY_COL = "ЮЛ"

MTS_CERT_COL = "Добавлен сертификат (МТС)"

# Булевые колонки: SOURCE → TARGET
COLS_SYNC = ["Добавлен сертификат", "Билеты продаются"]

# CF восстанавливаем на всех трёх
COLS_WITH_CF = ["Добавлен сертификат", MTS_CERT_COL, "Билеты продаются"]


# =======================
# FORWARD: СВОДНАЯ → Лист1 (bools only)
# =======================
def sync_bools_to_target(wb_src, wb_tgt) -> None:
    """Переносит Добавлен сертификат и Билеты продаются из СВОДНАЯ в Лист1."""
    if SRC_SHEET not in wb_src.sheetnames:
        raise RuntimeError(f'Source file: sheet "{SRC_SHEET}" not found')
    if TGT_SHEET not in wb_tgt.sheetnames:
        raise RuntimeError(f'Target file: sheet "{TGT_SHEET}" not found')

    ws_src = wb_src[SRC_SHEET]
    ws_tgt = wb_tgt[TGT_SHEET]

    src_map = header_index_map(ws_src)
    tgt_map = header_index_map(ws_tgt)

    if KEY_COL not in src_map:
        raise RuntimeError(f'Source sheet "{SRC_SHEET}": key column "{KEY_COL}" not found')

    # Ensure columns in TARGET
    for c in COLS_SYNC:
        if c not in tgt_map:
            h = last_header_col(ws_tgt) + 1
            ws_tgt.cell(row=1, column=h).value = c
    for c in COLS_WITH_CF:
        m = header_index_map(ws_tgt)
        if c not in m:
            h = last_header_col(ws_tgt) + 1
            ws_tgt.cell(row=1, column=h).value = c

    src_map = header_index_map(ws_src)
    tgt_map = header_index_map(ws_tgt)

    if KEY_COL not in tgt_map:
        raise RuntimeError(f'Target sheet "{TGT_SHEET}": key column "{KEY_COL}" not found')

    src_last = get_last_data_row(ws_src, src_map[KEY_COL], start_row=2)
    tgt_last = get_last_data_row(ws_tgt, tgt_map[KEY_COL], start_row=2)

    # Read SOURCE bools
    src_data: Dict[str, Dict[str, Optional[int]]] = {}
    for r in range(2, src_last + 1):
        key = get_cell_str(ws_src, r, src_map[KEY_COL])
        if not key:
            continue
        payload: Dict[str, Optional[int]] = {}
        for name in COLS_SYNC:
            if name not in src_map:
                payload[name] = None
                continue
            payload[name] = normalize_bool_to_01(ws_src.cell(row=r, column=src_map[name]).value)
        src_data[key] = payload

    # Existing TARGET rows
    tgt_row_by_key: Dict[str, int] = {}
    if tgt_last >= 2:
        for r in range(2, tgt_last + 1):
            key = get_cell_str(ws_tgt, r, tgt_map[KEY_COL])
            if key:
                tgt_row_by_key[key] = r

    template_row = 2 if ws_tgt.max_row >= 2 else 2
    max_style_col = last_header_col(ws_tgt)

    updated = 0
    inserted = 0
    append_row = tgt_last + 1 if tgt_last >= 2 else 2

    for key, payload in src_data.items():
        if key in tgt_row_by_key:
            rr = tgt_row_by_key[key]
            for name in COLS_SYNC:
                val = payload.get(name, None)
                if val is None:
                    continue
                ws_tgt.cell(row=rr, column=tgt_map[name]).value = val
            updated += 1
        else:
            rr = append_row
            append_row += 1

            if 2 <= template_row <= ws_tgt.max_row:
                copy_row_style(ws_tgt, template_row, rr, max_style_col)

            ws_tgt.cell(row=rr, column=tgt_map[KEY_COL]).value = key
            for name in COLS_SYNC:
                val = payload.get(name, None)
                ws_tgt.cell(row=rr, column=tgt_map[name]).value = val
            # MTS_CERT_COL для новых — 0 по умолчанию
            if MTS_CERT_COL in tgt_map:
                ws_tgt.cell(row=rr, column=tgt_map[MTS_CERT_COL]).value = 0
            inserted += 1

    # Нормализация
    new_tgt_last = max(get_last_data_row(ws_tgt, tgt_map[KEY_COL], start_row=2), 2)

    for name in COLS_WITH_CF:
        if name not in tgt_map:
            continue
        c = tgt_map[name]
        for r in range(2, new_tgt_last + 1):
            v = ws_tgt.cell(row=r, column=c).value
            if is_empty_cell(v):
                continue
            norm = normalize_bool_to_01(v)
            if norm is not None:
                ws_tgt.cell(row=r, column=c).value = norm

    # CF
    for name in COLS_WITH_CF:
        if name not in tgt_map:
            continue
        letter = col_to_letter(tgt_map[name])
        apply_bool_cf(ws_tgt, letter, start_row=2, end_row=new_tgt_last)

    print(f"  Forward (bools) done: updated={updated}, inserted={inserted}, total={len(src_data)}")


# =======================
# REVERSE: Лист1 → СВОДНАЯ (МТС cert only)
# =======================
def sync_mts_cert_back(wb_src, wb_tgt) -> None:
    """Переносит «Добавлен сертификат (МТС)» из Лист1 обратно в СВОДНАЯ."""
    ws_src = wb_src[SRC_SHEET]
    ws_tgt = wb_tgt[TGT_SHEET]

    src_map = header_index_map(ws_src)
    tgt_map = header_index_map(ws_tgt)

    if KEY_COL not in src_map or KEY_COL not in tgt_map:
        raise RuntimeError(f'Key column "{KEY_COL}" not found in both sheets')

    src_key_c = src_map[KEY_COL]
    tgt_key_c = tgt_map[KEY_COL]

    # Ensure col in SOURCE
    if MTS_CERT_COL not in src_map:
        ensure_columns_at_end(ws_src, [MTS_CERT_COL])
        src_map = header_index_map(ws_src)
    src_col_c = src_map[MTS_CERT_COL]

    if MTS_CERT_COL not in tgt_map:
        print("  Reverse skip: MTS cert column not found in TARGET")
        return
    tgt_col_c = tgt_map[MTS_CERT_COL]

    # Build dict from TARGET (Лист1)
    tgt_last = get_last_data_row(ws_tgt, tgt_key_c, start_row=2)
    data: Dict[str, int] = {}
    for r in range(2, tgt_last + 1):
        key = ws_tgt.cell(row=r, column=tgt_key_c).value
        if is_empty_cell(key):
            continue
        v = ws_tgt.cell(row=r, column=tgt_col_c).value
        norm = normalize_bool_to_01(v)
        if norm is None:
            continue
        data[str(key).strip()] = norm

    # Apply to SOURCE (СВОДНАЯ)
    src_last = get_last_data_row(ws_src, src_key_c, start_row=2)
    updated = 0
    for r in range(2, src_last + 1):
        key = ws_src.cell(row=r, column=src_key_c).value
        if is_empty_cell(key):
            continue
        k = str(key).strip()
        if k not in data:
            continue
        ws_src.cell(row=r, column=src_col_c).value = data[k]
        updated += 1

    # CF для всех трёх bool-колонок в СВОДНАЯ (иначе при сохранении слетают)
    bool_cols_src = ["Добавлен сертификат", MTS_CERT_COL, "Билеты продаются"]
    end = max(src_last, 2)
    for col_name in bool_cols_src:
        if col_name not in src_map:
            continue
        apply_bool_cf(ws_src, col_to_letter(src_map[col_name]), start_row=2, end_row=end)

    print(f"  Reverse (MTS cert) done: updated={updated}, keys_with_value={len(data)}")


# =======================
# ENTRYPOINT
# =======================
def main() -> None:
    print(f"Download SOURCE: {DISK_SOURCE_PATH}")
    src_bytes = disk_download(DISK_SOURCE_PATH, YANDEX_OAUTH_TOKEN)
    print(f"  SOURCE: {len(src_bytes)} bytes")

    print(f"Download TARGET: {DISK_TARGET_PATH}")
    tgt_bytes = disk_download(DISK_TARGET_PATH, YANDEX_OAUTH_TOKEN)
    print(f"  TARGET: {len(tgt_bytes)} bytes")

    wb_src = load_workbook(io.BytesIO(src_bytes))
    wb_tgt = load_workbook(io.BytesIO(tgt_bytes))

    # Forward: СВОДНАЯ → Лист1 (bools)
    print("Forward: СВОДНАЯ → Лист1 (bools)...")
    sync_bools_to_target(wb_src, wb_tgt)

    # Reverse: Лист1 → СВОДНАЯ (МТС cert)
    print("Reverse: Лист1 → СВОДНАЯ (МТС cert)...")
    sync_mts_cert_back(wb_src, wb_tgt)

    # Восстановить CF на листе «терминалы» (openpyxl теряет при save)
    SHEET_TERMINALS = "терминалы"
    if SHEET_TERMINALS in wb_tgt.sheetnames:
        ws_term = wb_tgt[SHEET_TERMINALS]
        term_map = header_index_map(ws_term)
        term_key = term_map.get("Агент ID (Столото)", term_map.get("ЮЛ"))
        term_last = get_last_data_row(ws_term, term_key, start_row=2) if term_key else 2
        for col_name in ["Добавлен сертификат", "Добавлен сертификат (МТС)"]:
            if col_name in term_map:
                apply_bool_cf(ws_term, col_to_letter(term_map[col_name]),
                              start_row=2, end_row=max(term_last, 2))
        print(f"  Restored CF on «{SHEET_TERMINALS}»")

    # Upload обоих файлов
    out_src = io.BytesIO()
    wb_src.save(out_src)
    out_tgt = io.BytesIO()
    wb_tgt.save(out_tgt)

    print(f"Upload SOURCE: {DISK_SOURCE_PATH}")
    disk_upload(DISK_SOURCE_PATH, out_src.getvalue(), YANDEX_OAUTH_TOKEN)

    print(f"Upload TARGET: {DISK_TARGET_PATH}")
    disk_upload(DISK_TARGET_PATH, out_tgt.getvalue(), YANDEX_OAUTH_TOKEN)

    print("Done: sheets sync")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
